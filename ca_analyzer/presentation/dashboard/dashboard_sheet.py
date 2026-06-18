from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ca_analyzer.presentation.dashboard.summary_cards import get_summary_cards_data
from ca_analyzer.presentation.dashboard.charts_data import write_dashboard_charts_data
from ca_analyzer.presentation.charts import (
    create_line_chart, create_stacked_bar_chart, create_combo_chart, 
    create_pie_chart, create_donut_chart
)
from ca_analyzer.presentation.styles import HDR_FILL, BORDER
from ca_analyzer.analytics.dashboard import generate_dashboard_kpis
from ca_analyzer.core.utilities import format_inr

def draw_kpi_card(ws, range_str: str, title: str, value: str, raw_value: float = None):
    """
    Draws a visual KPI card block on the worksheet:
    - Merges the top row for the Card Title (font size 9, bold, gray text)
    - Merges bottom rows for the Card Value (font size 13, bold, centered)
    - Applies thin gray border around the boundary
    - Net Cashflow card features dynamic soft green/red fill depending on cashflow
    """
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    
    # 1. Merge and style Top Row for Title
    ws.merge_cells(start_row=min_row, start_column=min_col, end_row=min_row, end_column=max_col)
    title_cell = ws.cell(row=min_row, column=min_col, value=title)
    title_cell.font = Font(name="Calibri", size=9, bold=True, color="595959")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 2. Merge and style Bottom Rows for Value
    ws.merge_cells(start_row=min_row + 1, start_column=min_col, end_row=max_row, end_column=max_col)
    val_cell = ws.cell(row=min_row + 1, column=min_col, value=value)
    
    # Determine fill and font color based on Net Cashflow sign
    if title == "Net Cashflow" and raw_value is not None:
        if raw_value >= 0:
            card_font = Font(name="Calibri", size=13, bold=True, color="375623")
            card_fill = PatternFill("solid", fgColor="E2EFDA")
        else:
            card_font = Font(name="Calibri", size=13, bold=True, color="C00000")
            card_fill = PatternFill("solid", fgColor="FCE4D6")
    else:
        card_font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
        card_fill = PatternFill("solid", fgColor="F2F5F9")
        
    val_cell.font = card_font
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply borders and backgrounds to all constituent cells in the card boundary
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = card_fill
            cell.border = BORDER

def create_dashboard_sheet(wb, df):
    """
    Constructs the premium visual Dashboard sheet in the Excel workbook,
    drawing a 3x3 grid of KPI cards and positioning 5 key charts.
    """
    ws = wb.create_sheet("📊 Executive Dashboard")
    ws.sheet_view.showGridLines = False
    
    # Title Block merged across columns A to Q (the grid space)
    ws.merge_cells("A1:Q1")
    title_cell = ws["A1"]
    title_cell.value = "FINANCIAL INTELLIGENCE EXECUTIVE DASHBOARD"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, 18):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER
        
    # Spacer row 2
    ws.row_dimensions[2].height = 15
    for col in range(1, 18):
        ws.cell(row=2, column=col).border = BORDER
        
    # Fetch KPIs values
    kpis = generate_dashboard_kpis(df)
    
    def format_to_lakhs(val):
        lakhs = val / 100000.0
        return f"₹ {lakhs:.2f} L"
        
    # Standard formatters
    credits_val = format_to_lakhs(kpis.get("Total Credits", 0))
    debits_val = format_to_lakhs(kpis.get("Total Debits", 0))
    cashflow_val = format_to_lakhs(kpis.get("Net Cashflow", 0))
    hibal_val = format_to_lakhs(kpis.get("Highest Balance", 0))
    lobal_val = format_to_lakhs(kpis.get("Lowest Balance", 0))
    banks_count = str(kpis.get("Number of Banks", 0))
    txns_count = f"{kpis.get('Total Transactions', 0):,}"
    inflow_val = format_to_lakhs(kpis.get("Average Monthly Inflow", 0))
    outflow_val = format_to_lakhs(kpis.get("Average Monthly Outflow", 0))
    
    # 3x3 Card layout positioning specification
    # (Range, Title, formatted_value, raw_value)
    KPI_LAYOUT = [
        ("B4:C6", "Total Credits", credits_val, kpis.get("Total Credits")),
        ("E4:F6", "Total Debits", debits_val, kpis.get("Total Debits")),
        ("H4:I6", "Net Cashflow", cashflow_val, kpis.get("Net Cashflow")),
        
        ("B8:C10", "Highest Balance", hibal_val, kpis.get("Highest Balance")),
        ("E8:F10", "Lowest Balance", lobal_val, kpis.get("Lowest Balance")),
        ("H8:I10", "Number of Banks", banks_count, kpis.get("Number of Banks")),
        
        ("B12:C14", "Total Transactions", txns_count, kpis.get("Total Transactions")),
        ("E12:F14", "Average Monthly Inflow", inflow_val, kpis.get("Average Monthly Inflow")),
        ("H12:I14", "Average Monthly Outflow", outflow_val, kpis.get("Average Monthly Outflow"))
    ]
    
    # Set standard column widths for dashboard spacing
    ws.column_dimensions["A"].width = 3   # Left margin
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 3   # Spacer
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 3   # Spacer
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 3   # Spacer
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 15
    ws.column_dimensions["M"].width = 15
    ws.column_dimensions["N"].width = 15
    
    # Draw KPI cards
    for range_str, title, value, raw in KPI_LAYOUT:
        draw_kpi_card(ws, range_str, title, value, raw)
        
    # Write charts data to hidden helper columns Z+
    refs = write_dashboard_charts_data(ws, df)
    
    # Add charts starting at row 16
    # 1. Credit vs Debit Trend Line Chart at B16
    c_trend = create_line_chart(
        "Credit vs Debit Trend", ws,
        min_col=refs["trends"]["data_min_col"],
        min_row=1,
        max_col=refs["trends"]["data_max_col"],
        max_row=refs["trends"]["data_max_row"],
        cats_col=refs["trends"]["cats_col"],
        cats_row_start=2,
        cats_row_end=refs["trends"]["cats_row_end"]
    )
    ws.add_chart(c_trend, "B16")
    
    # 2. Monthly Closing Balance Combo (Column + Line) at J16
    c_bal = create_combo_chart(
        "Monthly Closing Balance", ws,
        min_col=refs["balance"]["data_min_col"],
        min_row=1,
        max_col=refs["balance"]["data_max_col"],
        max_row=refs["balance"]["data_max_row"],
        cats_col=refs["balance"]["cats_col"],
        cats_row_start=2,
        cats_row_end=refs["balance"]["cats_row_end"]
    )
    ws.add_chart(c_bal, "K16")
    
    # 3. Expense Category Pie Chart at B34
    c_exp = create_pie_chart(
        "Expense Category Contribution", ws,
        min_col=refs["expense"]["data_min_col"],
        min_row=1,
        max_col=refs["expense"]["data_max_col"],
        max_row=refs["expense"]["data_max_row"],
        cats_col=refs["expense"]["cats_col"],
        cats_row_start=2,
        cats_row_end=refs["expense"]["cats_row_end"]
    )
    ws.add_chart(c_exp, "B34")
    
    # 4. Bank Contribution Donut Chart at J34
    c_bank = create_donut_chart(
        "Bank Inflow Contribution", ws,
        min_col=refs["bank"]["data_min_col"],
        min_row=1,
        max_col=refs["bank"]["data_max_col"],
        max_row=refs["bank"]["data_max_row"],
        cats_col=refs["bank"]["cats_col"],
        cats_row_start=2,
        cats_row_end=refs["bank"]["cats_row_end"]
    )
    ws.add_chart(c_bank, "K34")
    
    # 5. Monthly Cashflow Stacked Column Chart at B52
    c_flow = create_stacked_bar_chart(
        "Monthly Cashflow (Credit Inflows by Bank)", ws,
        min_col=refs["cashflow"]["data_min_col"],
        min_row=1,
        max_col=refs["cashflow"]["data_max_col"],
        max_row=refs["cashflow"]["data_max_row"],
        cats_col=refs["cashflow"]["cats_col"],
        cats_row_start=2,
        cats_row_end=refs["cashflow"]["cats_row_end"]
    )
    ws.add_chart(c_flow, "B52")
    
    # Hide all helper data columns Z through AZ (indexes 26 to 52)
    for col_idx in range(26, 53):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].hidden = True
        
    # Freezing panes to none (A1) for dashboard
    ws.freeze_panes = "A1"
