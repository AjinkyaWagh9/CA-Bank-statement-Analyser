import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from ca_analyzer.presentation.dashboard import create_dashboard_sheet
from ca_analyzer.presentation.workbook_formatter import populate_and_format_table_sheet, format_custom_table_range
from ca_analyzer.presentation.styles import HDR_FILL, HDR_FONT, ALT_FILL, BORDER, CENTER, CURRENCY_FORMAT
from ca_analyzer.presentation import formula_refs
from ca_analyzer.core.utilities import format_inr
from ca_analyzer.core.schemas import LOCKED_COLUMNS, UNLOCKED_COLUMNS
from ca_analyzer.analytics import (
    generate_income_analysis, generate_expense_analysis, generate_cash_deposit_analysis,
    generate_high_value_transactions, generate_loan_analysis, generate_gst_analysis,
    generate_tds_analysis, generate_investment_analysis, generate_risk_flags
)
from ca_analyzer.validation.reconciliation import reconcile_all
from ca_analyzer.transaction_engine.loan_matcher import build_loan_ledger

# ---------------------------------------------------------------------------
# Canonical category lists (must match category_rules.yaml output exactly)
# ---------------------------------------------------------------------------
INCOME_CATEGORIES = [
    "Salary", "House Property", "Other Sources", "Business/Profession",
    "Capital Gains", "Loans", "Transfers", "Miscellaneous",
]

EXPENSE_CATEGORIES = [
    "Loans", "Insurance", "Taxes", "Investments", "Utilities", "Rent",
    "Cash", "Bounces", "Food", "Travel", "Shopping", "Medical", "Others",
]

# ---------------------------------------------------------------------------
# Shared helper for writing formula-based summary sheets
# ---------------------------------------------------------------------------

def _write_formula_sheet_title(ws, title_text: str, num_cols: int):
    """Writes row 1 (merged title) and row 2 (spacer) for a formula sheet."""
    ws.sheet_view.showGridLines = False
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER

    ws.row_dimensions[2].height = 15
    for col in range(1, num_cols + 1):
        ws.cell(row=2, column=col).border = BORDER


def _write_header_row(ws, row: int, headers: list):
    """Writes styled headers at the given row."""
    ws.row_dimensions[row].height = 22
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_total_row(ws, row: int, num_cols: int):
    """Makes the TOTAL row bold with a top border."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", bold=True)
        cell.border = BORDER


def _apply_currency_format(ws, start_row: int, end_row: int, col_indices: list):
    """Applies ₹ currency format to a list of column indices (1-based) in a row range."""
    for row in range(start_row, end_row + 1):
        for col in col_indices:
            cell = ws.cell(row=row, column=col)
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = Alignment(horizontal="right", vertical="center")

def build_ca_report(df: pd.DataFrame, output_path: str) -> str:
    """
    Main pipeline entry point for presenting the workbook.
    Accepts consolidated normalized dataframe and writes to Consolidated_CA_Analysis.xlsx.
    """
    wb = openpyxl.Workbook()
    
    # Apply workbook properties metadata
    wb.properties.creator = "CA Financial Intelligence Analyzer"
    wb.properties.title = "Consolidated Bank Statement Analysis"
    wb.properties.subject = "Multi-Bank Financial Analysis Report"
    wb.properties.description = "CA / ITR Filing Consolidated Bank Statement Report"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    # 1. 📋 Cover Sheet
    create_cover_sheet(wb, df)
    
    # 2. Executive Summary (observations, compliance metrics, and checklist)
    create_executive_summary_sheet(wb, df)
    
    # 4. Bank Wise Summary
    create_bank_wise_summary_sheet(wb, df)
    
    # 5. Monthly Cashflow
    create_monthly_cashflow_sheet(wb, df)
    
    # 6. Income Analysis
    create_income_analysis_sheet(wb, df)
    
    # 7. Expense Analysis
    create_expense_analysis_sheet(wb, df)
    
    # 8. Cash Deposit Analysis
    create_cash_deposit_sheet(wb, df)
    
    # 9. High Value Transactions
    create_high_value_sheet(wb, df)
    
    # 10. Loan Analysis
    create_loan_sheet(wb, df)
    
    # 11. GST Analysis
    create_gst_sheet(wb, df)
    
    # 12. TDS Analysis
    create_tds_sheet(wb, df)
    
    # 13. Investment Analysis
    create_investment_sheet(wb, df)
    
    # 14. Risk Flags
    create_risk_flags_sheet(wb, df)

    # 15. Tax Payments
    create_tax_payments_sheet(wb, df)

    # 16. 80C Deduction Tracker
    create_80c_tracker_sheet(wb, df)

    # 17. Capital Gain
    create_capital_gain_sheet(wb, df)

    # 18. Drawings
    create_drawings_sheet(wb, df)

    # 19. Unclassified Transactions
    create_unclassified_sheet(wb, df)

    # 20. Run reconciliation and pass discrepancies to CA Observations
    reconciliation_discrepancies = reconcile_all(df)

    # 21. CA Observations
    create_ca_observations_sheet(wb, df, reconciliation_discrepancies)

    # 22. 🔁 Inter-Bank Transfers sheet
    create_inter_bank_transfers_sheet(wb, df)

    # 23. 🏦 Loan Ledger sheet
    create_loan_ledger_sheet(wb, df)

    # 24. 📉 EMI Analysis sheet
    create_emi_analysis_sheet(wb, df)

    # 25. 👥 Related Party sheet
    create_related_party_sheet(wb, df)

    # 26+. Bank-wise transaction ledgers
    for bank in sorted(df["Bank_Name"].unique()):
        bank_df = df[df["Bank_Name"] == bank].copy()
        create_transaction_sheet(wb, f"📒 {bank} Transactions", bank_df)

    # 16. Consolidated All Transactions master sheet (Phase 0 — editable overrides + locked evidence)
    # Returns col_name_to_idx so named ranges can reference exact column letters.
    col_name_to_idx = create_consolidated_master_sheet(wb, df)

    # 17. Phase 1.1 — define workbook-level named ranges pointing at master sheet columns
    formula_refs.define_named_ranges(wb, col_name_to_idx)

    wb.save(output_path)
    return output_path

def generate_monthly_cashflow_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """Creates a monthly credit inflow pivot table by bank."""
    df = df.copy()
    df["Month"] = df["Date"].dt.strftime("%Y-%m")
    
    credits = df[df["Credit"] > 0.0]
    if credits.empty:
        return pd.DataFrame(columns=["Month", "Total"])
        
    pivot = credits.pivot_table(
        index="Month",
        columns="Bank_Name",
        values="Credit",
        aggfunc="sum",
        fill_value=0.0
    ).reset_index()
    
    bank_cols = [c for c in pivot.columns if c != "Month"]
    pivot["Total"] = pivot[bank_cols].sum(axis=1)
    return pivot

def create_cover_sheet(wb, df):
    """Creates a premium metadata Cover Sheet as the entrypoint worksheet."""
    ws = wb.create_sheet("📋 Cover Sheet")
    ws.sheet_view.showGridLines = False
    
    # Merged title block across columns A to E
    ws.merge_cells("A1:E2")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 25
    
    title_cell = ws["A1"]
    title_cell.value = "FINANCIAL INTELLIGENCE CONSOLIDATED REPORT"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for r in [1, 2]:
        for col in range(1, 6):
            c = ws.cell(row=r, column=col)
            c.fill = HDR_FILL
            c.border = BORDER
            
    # Spacer rows 3 and 4
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = BORDER
        ws.cell(row=4, column=col).border = BORDER
        
    # Table headers in Row 5
    ws["B5"] = "Metadata Field"
    ws["C5"] = "Analysis Properties"
    for col in ["B5", "C5"]:
        ws[col].fill = HDR_FILL
        ws[col].font = HDR_FONT
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
        ws[col].border = BORDER
    ws.row_dimensions[5].height = 22
    
    holder_name = df["Person_Name"].iloc[0] if not df.empty else "N/A"
    start_date = df["Statement_Start"].iloc[0] if not df.empty else "N/A"
    end_date = df["Statement_End"].iloc[0] if not df.empty else "N/A"
    banks_list = sorted(list(df["Bank_Name"].unique()))
    banks_str = ", ".join(banks_list)
    
    total_credits = df["Credit"].sum()
    total_debits = df["Debit"].sum()
    net_cashflow = total_credits - total_debits
    
    import datetime
    date_str = datetime.date.today().strftime("%d-%m-%Y")
    
    metadata_rows = [
        ("Client Account Holder", holder_name),
        ("Statement Period", f"{start_date} to {end_date}"),
        ("Audited Bank Accounts", f"{len(banks_list)} ({banks_str})"),
        ("Consolidated Inflows (Credits)", format_inr(total_credits)),
        ("Consolidated Outflows (Debits)", format_inr(total_debits)),
        ("Net In-file Cashflow", format_inr(net_cashflow)),
        ("Date Generated", date_str),
        ("Report Engine", "CA Financial Intelligence Analyzer v2.1")
    ]
    
    for r, (field, value) in enumerate(metadata_rows, start=6):
        ws[f"B{r}"] = field
        ws[f"C{r}"] = value
        
        ws[f"B{r}"].border = BORDER
        ws[f"C{r}"].border = BORDER
        ws[f"B{r}"].font = Font(name="Calibri", bold=True)
        ws[f"C{r}"].fill = ALT_FILL
        ws.row_dimensions[r].height = 18
        
        # Format C row values as currencies where applicable
        if any(x in field.lower() for x in ["inflow", "outflow", "cashflow"]):
            ws[f"C{r}"].number_format = "₹#,##0.00"
            ws[f"C{r}"].alignment = Alignment(horizontal="right")
            try:
                ws[f"C{r}"].value = float(str(value).replace("₹", "").replace(",", "").strip())
            except ValueError:
                pass
                
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    
    ws.freeze_panes = "A1"

def create_executive_summary_sheet(wb, df):
    ws = wb.create_sheet("📝 Executive Summary")
    ws.sheet_view.showGridLines = False
    
    # Merged title row
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value="EXECUTIVE CA SUMMARY & COMPLIANCE AUDIT")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, 4):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER
        
    # Spacer
    ws.row_dimensions[2].height = 15
    for col in range(1, 4):
        ws.cell(row=2, column=col).border = BORDER
        
    # Write metadata summary block
    ws["A3"] = "Metric"
    ws["B3"] = "Details"
    ws["A3"].fill = HDR_FILL
    ws["A3"].font = HDR_FONT
    ws["A3"].border = BORDER
    ws["B3"].fill = HDR_FILL
    ws["B3"].font = HDR_FONT
    ws["B3"].border = BORDER
    ws.row_dimensions[3].height = 22
    
    holder_name = df["Person_Name"].iloc[0] if not df.empty else "N/A"
    start_date = df["Statement_Start"].iloc[0] if not df.empty else "N/A"
    end_date = df["Statement_End"].iloc[0] if not df.empty else "N/A"
    
    summary_rows = [
        ("Client Account Holder", holder_name),
        ("Statement Period Covered", f"{start_date} to {end_date}"),
        ("Total Transaction Count", len(df)),
        ("Consolidated Accounts Audited", df["Bank_Name"].nunique())
    ]
    for r, (metric, details) in enumerate(summary_rows, start=4):
        ws[f"A{r}"] = metric
        ws[f"B{r}"] = details
        ws[f"A{r}"].border = BORDER
        ws[f"B{r}"].border = BORDER
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"].fill = ALT_FILL
        ws.row_dimensions[r].height = 18
        
    # Spacer in row 8
    ws.row_dimensions[8].height = 15
    for col in range(1, 4):
        ws.cell(row=8, column=col).border = BORDER
        
    # Write Compliance Checklist
    ws["A9"] = "Audit Checklist Rule"
    ws["B9"] = "Status"
    ws["C9"] = "Observations & Recommendations"
    for col in ["A9", "B9", "C9"]:
        ws[col].fill = HDR_FILL
        ws[col].font = HDR_FONT
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
        ws[col].border = BORDER
    ws.row_dimensions[9].height = 22
    
    # Perform checklist audits
    checklist = []
    # Check 1: SFT Cash Limit
    total_cash_dep = df[(df["Credit"] > 0.0) & (df["Category"] == "Cash")]["Credit"].sum()
    if total_cash_dep > 200000.0:
        checklist.append(("SFT Savings Account Cash Deposit Limit (₹2 Lakh)", "🚩 RED FLAG", f"Total cash deposits of {format_inr(total_cash_dep)} exceed the ₹2 Lakh SFT threshold. Reportable under Rule 114E."))
    else:
        checklist.append(("SFT Savings Account Cash Deposit Limit (₹2 Lakh)", "✅ PASS", f"Total cash deposits of {format_inr(total_cash_dep)} are within safe SFT limits."))
        
    # Check 2: High Value Transactions
    large_txns_count = df[(df["Credit"] >= 50000.0) | (df["Debit"] >= 50000.0)].shape[0]
    if large_txns_count > 0:
        checklist.append(("High Value Transactions (Sec 285BA / AIS Matching)", "⚠️ WARNING", f"Found {large_txns_count} transaction(s) >= ₹50,000. Ensure verification against AIS details."))
    else:
        checklist.append(("High Value Transactions (Sec 285BA / AIS Matching)", "✅ PASS", "No transactions >= ₹50,000 detected."))
        
    # Check 3: Chapter VI-A Investment Deductions (Sec 80C / 80D)
    has_invest = df[df["Category"] == "Investments"].shape[0] > 0
    if has_invest:
        checklist.append(("Tax Saving Investment Audit (Sec 80C / 80D)", "✅ DETECTED", "Deduction flows found (SIP/PPF/NPS/Insurance). Verify investment challans for final computations."))
    else:
        checklist.append(("Tax Saving Investment Audit (Sec 80C / 80D)", "⚠️ NOT FOUND", "No common tax saving investments (Sec 80C/80D) were detected in the statements."))
        
    # Check 4: Cheque Bounces
    bounce_count = df[df["Category"] == "Bounces"].shape[0]
    if bounce_count > 0:
        checklist.append(("Cheque Bounce & ECS Return Audit (Sec 138 NI Act)", "🚩 RED FLAG", f"Detected {bounce_count} bounce/ECS return events. Review customer ledger balances."))
    else:
        checklist.append(("Cheque Bounce & ECS Return Audit (Sec 138 NI Act)", "✅ PASS", "No cheque bounces or mandate returns detected."))
        
    for i, (rule, status, obs) in enumerate(checklist, start=10):
        ws[f"A{i}"] = rule
        ws[f"B{i}"] = status
        ws[f"C{i}"] = obs
        ws.row_dimensions[i].height = 18
        for col in ["A", "B", "C"]:
            cell = ws[f"{col}{i}"]
            cell.border = BORDER
            if col == "B":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if "RED FLAG" in status:
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    cell.font = Font(name="Calibri", bold=True, color="C00000")
                elif "WARNING" in status or "NOT FOUND" in status:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    cell.font = Font(name="Calibri", bold=True, color="7F6000")
                else:
                    cell.fill = PatternFill("solid", fgColor="E2EFDA")
                    cell.font = Font(name="Calibri", bold=True, color="375623")
            elif i % 2 == 0:
                cell.fill = ALT_FILL
                
    # Auto column widths for Executive Summary
    for col in ["A", "B", "C"]:
        max_len = max(len(str(ws[f"{col}{r}"].value or "")) for r in range(3, 14))
        ws.column_dimensions[col].width = max(max_len + 3, 15)
        
    # Set freeze panes to B4 (to keep summary headers visible)
    ws.freeze_panes = "B4"

def create_bank_wise_summary_sheet(wb, df):
    """
    Phase 1.1 — Formula version of the Bank Wise Summary sheet.

    Credits and transaction counts use SUMIFS/COUNTIFS against the m_Credit / m_Bank
    named ranges so that CA edits to the master sheet flow through automatically.

    Closing Balance is computed with Python (requires sorting — not feasible as a
    single SUMIFS formula) and written as a static value with a note.
    """
    ws = wb.create_sheet("🏦 Bank Wise Summary")

    banks = list(df["Bank_Name"].unique())
    headers = ["Bank", "Credits", "Debits", "Closing Balance", "Transaction Count"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "🏦 BANK WISE STATEMENT SUMMARY", num_cols)
    _write_header_row(ws, 3, headers)

    first_data_row = 4
    for i, bank in enumerate(banks):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18

        # Col A: bank label (static)
        ws.cell(row=row, column=1, value=bank).border = BORDER

        # Col B: Credits via SUMIFS formula
        cell_b = ws.cell(row=row, column=2, value=f"=SUMIFS(m_Credit,m_Bank,A{row})")
        cell_b.border = BORDER

        # Col C: Debits via SUMIFS formula
        cell_c = ws.cell(row=row, column=3, value=f"=SUMIFS(m_Debit,m_Bank,A{row})")
        cell_c.border = BORDER

        # Col D: Closing Balance — Python-computed (sort required; not expressible as SUMIFS)
        bank_df = df[df["Bank_Name"] == bank].copy()
        last_bal = (
            bank_df.sort_values(by=["Date", "Balance"])["Balance"].iloc[-1]
            if not bank_df.empty else 0.0
        )
        cell_d = ws.cell(row=row, column=4, value=last_bal)
        cell_d.border = BORDER

        # Col E: Transaction Count via COUNTIFS formula
        cell_e = ws.cell(row=row, column=5, value=f"=COUNTIFS(m_Bank,A{row})")
        cell_e.border = BORDER

    # TOTAL row
    last_data_row = first_data_row + len(banks) - 1
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(row=total_row, column=1, value="TOTAL").border = BORDER
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})").border = BORDER
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})").border = BORDER
    ws.cell(row=total_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})").border = BORDER
    ws.cell(row=total_row, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})").border = BORDER
    _style_total_row(ws, total_row, num_cols)

    # Apply currency format to Credits, Debits, Closing Balance columns (B, C, D = 2, 3, 4)
    _apply_currency_format(ws, first_data_row, total_row, [2, 3, 4])

    # Column widths
    ws.column_dimensions["A"].width = 25
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions["E"].width = 20

    ws.freeze_panes = "B4"

def create_monthly_cashflow_sheet(wb, df):
    """
    Phase 1.1 — Formula version of the Monthly Cashflow sheet.

    Each cell uses SUMIFS with DATE/EOMONTH to aggregate credit inflows for the
    month stored as a 'YYYY-MM' label in column A.  Bank columns come from the
    m_Credit / m_Bank / m_Date named ranges, so CA edits flow through automatically.
    """
    ws = wb.create_sheet("📅 Monthly Cashflow")

    # Compute unique months from the dataframe (Python side — labels only)
    df_copy = df.copy()
    df_copy["Month"] = df_copy["Date"].dt.strftime("%Y-%m")
    months = sorted(df_copy["Month"].dropna().unique().tolist())

    banks = list(df["Bank_Name"].unique())
    headers = ["Month"] + banks + ["Total"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "📅 MONTHLY CASHFLOW ANALYSIS (CREDIT INFLOWS)", num_cols)
    _write_header_row(ws, 3, headers)

    # Write bank names into header row so mixed-refs B$3 etc. resolve correctly
    # (already done by _write_header_row above)

    first_data_row = 4
    last_bank_col = 1 + len(banks)  # 1-based col index of last bank column

    for i, month in enumerate(months):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18

        # Col A: month label (static)
        ws.cell(row=row, column=1, value=month).border = BORDER

        # Bank columns: SUMIFS using DATE/EOMONTH, reading year/month from A{row}
        for b_idx, _ in enumerate(banks):
            col = 2 + b_idx  # column B, C, ...
            formula = (
                f"=SUMIFS(m_Credit,m_Bank,{get_column_letter(col)}$3,"
                f"m_Date,\">=\"&DATE(LEFT(A{row},4)*1,MID(A{row},6,2)*1,1),"
                f"m_Date,\"<=\"&EOMONTH(DATE(LEFT(A{row},4)*1,MID(A{row},6,2)*1,1),0))"
            )
            cell = ws.cell(row=row, column=col, value=formula)
            cell.border = BORDER

        # Total column (last col): SUM across bank columns in this row
        total_col = 1 + len(banks) + 1
        first_bank_letter = get_column_letter(2)
        last_bank_letter = get_column_letter(last_bank_col)
        total_cell = ws.cell(
            row=row, column=total_col,
            value=f"=SUM({first_bank_letter}{row}:{last_bank_letter}{row})"
        )
        total_cell.border = BORDER

    # TOTAL row
    last_data_row = first_data_row + len(months) - 1
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(row=total_row, column=1, value="TOTAL").border = BORDER

    for b_idx in range(len(banks)):
        col = 2 + b_idx
        col_letter = get_column_letter(col)
        ws.cell(
            row=total_row, column=col,
            value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        ).border = BORDER

    total_col = 1 + len(banks) + 1
    total_col_letter = get_column_letter(total_col)
    ws.cell(
        row=total_row, column=total_col,
        value=f"=SUM({total_col_letter}{first_data_row}:{total_col_letter}{last_data_row})"
    ).border = BORDER
    _style_total_row(ws, total_row, num_cols)

    # Currency format for all numeric columns (2 through num_cols)
    _apply_currency_format(ws, first_data_row, total_row, list(range(2, num_cols + 1)))

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.freeze_panes = "B4"

def create_income_analysis_sheet(wb, df):
    """
    Phase 1.1 — Formula version of the Income Analysis sheet.

    Category × Bank matrix where each cell uses:
      =SUMIFS(m_Credit, m_CatFinal, $A{row}, m_Bank, {BankCol}$3)
    so that CA overrides in Category_Final flow through automatically.
    """
    ws = wb.create_sheet("💰 Income Analysis")

    banks = list(df["Bank_Name"].unique())
    categories = INCOME_CATEGORIES
    headers = ["Category"] + banks + ["Total"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "💰 INCOME SOURCE ANALYSIS BY ACCOUNT", num_cols)
    _write_header_row(ws, 3, headers)

    first_data_row = 4
    last_bank_col = 1 + len(banks)  # 1-based col index of the last bank column

    for i, cat in enumerate(categories):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18

        # Col A: category label (static, anchored with $ for mixed-ref formulas)
        ws.cell(row=row, column=1, value=cat).border = BORDER

        # Bank columns: SUMIFS matching Category_Final ($A{row}) and Bank (header row 3)
        for b_idx in range(len(banks)):
            col = 2 + b_idx
            col_letter = get_column_letter(col)
            formula = f"=SUMIFS(m_Credit,m_CatFinal,$A{row},m_Bank,{col_letter}$3)"
            ws.cell(row=row, column=col, value=formula).border = BORDER

        # Total column: SUM across bank columns
        total_col = last_bank_col + 1
        first_bank_letter = get_column_letter(2)
        last_bank_letter = get_column_letter(last_bank_col)
        ws.cell(
            row=row, column=total_col,
            value=f"=SUM({first_bank_letter}{row}:{last_bank_letter}{row})"
        ).border = BORDER

    # TOTAL row
    last_data_row = first_data_row + len(categories) - 1
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(row=total_row, column=1, value="TOTAL").border = BORDER

    for b_idx in range(len(banks)):
        col = 2 + b_idx
        col_letter = get_column_letter(col)
        ws.cell(
            row=total_row, column=col,
            value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        ).border = BORDER

    total_col = last_bank_col + 1
    total_col_letter = get_column_letter(total_col)
    ws.cell(
        row=total_row, column=total_col,
        value=f"=SUM({total_col_letter}{first_data_row}:{total_col_letter}{last_data_row})"
    ).border = BORDER
    _style_total_row(ws, total_row, num_cols)

    # Currency format for all numeric columns
    _apply_currency_format(ws, first_data_row, total_row, list(range(2, num_cols + 1)))

    # Column widths
    ws.column_dimensions["A"].width = 25
    for col in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.freeze_panes = "B4"

def create_expense_analysis_sheet(wb, df):
    """
    Phase 1.1 — Formula version of the Expense Analysis sheet.

    Category × Bank matrix (debit side) where each cell uses:
      =SUMIFS(m_Debit, m_CatFinal, $A{row}, m_Bank, {BankCol}$3)
    so that CA overrides in Category_Final flow through automatically.
    """
    ws = wb.create_sheet("💸 Expense Analysis")

    banks = list(df["Bank_Name"].unique())
    categories = EXPENSE_CATEGORIES
    headers = ["Category"] + banks + ["Total"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "💸 EXPENSE DISTRIBUTION BY ACCOUNT", num_cols)
    _write_header_row(ws, 3, headers)

    first_data_row = 4
    last_bank_col = 1 + len(banks)

    for i, cat in enumerate(categories):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18

        # Col A: category label (static)
        ws.cell(row=row, column=1, value=cat).border = BORDER

        # Bank columns: SUMIFS on debit side matching Category_Final and Bank
        for b_idx in range(len(banks)):
            col = 2 + b_idx
            col_letter = get_column_letter(col)
            formula = f"=SUMIFS(m_Debit,m_CatFinal,$A{row},m_Bank,{col_letter}$3)"
            ws.cell(row=row, column=col, value=formula).border = BORDER

        # Total column: SUM across bank columns
        total_col = last_bank_col + 1
        first_bank_letter = get_column_letter(2)
        last_bank_letter = get_column_letter(last_bank_col)
        ws.cell(
            row=row, column=total_col,
            value=f"=SUM({first_bank_letter}{row}:{last_bank_letter}{row})"
        ).border = BORDER

    # TOTAL row
    last_data_row = first_data_row + len(categories) - 1
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(row=total_row, column=1, value="TOTAL").border = BORDER

    for b_idx in range(len(banks)):
        col = 2 + b_idx
        col_letter = get_column_letter(col)
        ws.cell(
            row=total_row, column=col,
            value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        ).border = BORDER

    total_col = last_bank_col + 1
    total_col_letter = get_column_letter(total_col)
    ws.cell(
        row=total_row, column=total_col,
        value=f"=SUM({total_col_letter}{first_data_row}:{total_col_letter}{last_data_row})"
    ).border = BORDER
    _style_total_row(ws, total_row, num_cols)

    # Currency format for all numeric columns
    _apply_currency_format(ws, first_data_row, total_row, list(range(2, num_cols + 1)))

    # Column widths
    ws.column_dimensions["A"].width = 25
    for col in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.freeze_panes = "B4"

def create_cash_deposit_sheet(wb, df):
    ws = wb.create_sheet("💵 Cash Deposit Analysis")
    res = generate_cash_deposit_analysis(df)
    summary = res["summary"]
    flagged = res["flagged"]
    
    # 1. Merged Title
    ws.sheet_view.showGridLines = False
    max_cols = max(summary.shape[1], flagged.shape[1], 1)
    last_col_letter = get_column_letter(max_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=1, value="💵 CASH INFLOW & SFT AUDIT OBSERVATIONS")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, max_cols + 1):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER
        
    # 2. Blank Spacer in Row 2
    ws.row_dimensions[2].height = 15
    for col in range(1, max_cols + 1):
        ws.cell(row=2, column=col).border = BORDER
        
    # 3. Write Summary Table starting at Row 3 (headers)
    ws.cell(row=3, column=1, value="Bank")
    ws.cell(row=3, column=2, value="Total Cash Deposits")
    
    for r_idx, (_, row) in enumerate(summary.iterrows(), start=4):
        ws.cell(row=r_idx, column=1, value=row["Bank"])
        ws.cell(row=r_idx, column=2, value=row["Total Cash Deposits"])
        
    # Format Summary Table
    format_custom_table_range(
        ws, 
        start_row=4, 
        end_row=3 + len(summary), 
        max_col=summary.shape[1], 
        table_title_clean="Cash_Summary", 
        header_row=3,
        apply_table=True
    )
    
    # 4. Spacer Rows and Flagged Table Sub-Header
    spacer_row = 3 + len(summary) + 1
    ws.row_dimensions[spacer_row].height = 15
    for col in range(1, max_cols + 1):
        ws.cell(row=spacer_row, column=col).border = BORDER
        
    sub_header_row = spacer_row + 1
    ws.merge_cells(start_row=sub_header_row, start_column=1, end_row=sub_header_row, end_column=max_cols)
    sub_cell = ws.cell(row=sub_header_row, column=1, value="🚩 FLAGGED CASH DEPOSITS FOR SCRUTINY")
    sub_cell.font = Font(name="Calibri", bold=True, size=11, color="C00000")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[sub_header_row].height = 20
    
    for col in range(1, max_cols + 1):
        ws.cell(row=sub_header_row, column=col).border = BORDER
        
    spacer_row_2 = sub_header_row + 1
    ws.row_dimensions[spacer_row_2].height = 10
    for col in range(1, max_cols + 1):
        ws.cell(row=spacer_row_2, column=col).border = BORDER
        
    flagged_header_row = spacer_row_2 + 1
    for col_idx, col_name in enumerate(flagged.columns, start=1):
        ws.cell(row=flagged_header_row, column=col_idx, value=col_name)
        
    for r_idx, (_, row) in enumerate(flagged.iterrows(), start=flagged_header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            if pd.isna(val):
                ws.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)
                
    # Format Flagged Table
    format_custom_table_range(
        ws, 
        start_row=flagged_header_row + 1, 
        end_row=flagged_header_row + len(flagged), 
        max_col=flagged.shape[1], 
        table_title_clean="Cash_Flagged", 
        header_row=flagged_header_row,
        apply_table=False
    )
    
    ws.freeze_panes = "B4"

def create_high_value_sheet(wb, df):
    ws = wb.create_sheet("🚨 High Value Transactions")
    hvt_df = generate_high_value_transactions(df)
    populate_and_format_table_sheet(ws, "🚨 HIGH VALUE TRANSACTIONS AUDIT REPORT (>= ₹50K)", hvt_df)

def create_loan_sheet(wb, df):
    ws = wb.create_sheet("🏠 Loan Analysis")
    loan_df = generate_loan_analysis(df)
    populate_and_format_table_sheet(ws, "🏠 EMI & LIABILITY REPAYMENT REPORT", loan_df)

def create_gst_sheet(wb, df):
    ws = wb.create_sheet("🧾 GST Analysis")
    gst_df = generate_gst_analysis(df)
    populate_and_format_table_sheet(ws, "🧾 GST PAYMENTS, REFUNDS, & RECEIPTS", gst_df)

def create_tds_sheet(wb, df):
    ws = wb.create_sheet("🧾 TDS Analysis")
    tds_df = generate_tds_analysis(df)
    populate_and_format_table_sheet(ws, "🧾 TDS APPLICABILITY & AUDIT COMPLIANCE", tds_df)

def create_investment_sheet(wb, df):
    ws = wb.create_sheet("📈 Investment Analysis")
    inv_df = generate_investment_analysis(df)
    populate_and_format_table_sheet(ws, "📈 PORTFOLIO OUTFLOWS & TAX SAVING INVESTMENTS", inv_df)

def create_risk_flags_sheet(wb, df):
    ws = wb.create_sheet("⚠️ Risk Flags")
    risk_df = generate_risk_flags(df)
    populate_and_format_table_sheet(ws, "⚠️ SUSPICIOUS PATTERNS & REGULATORY AUDIT FLAGS", risk_df)

def create_tax_payments_sheet(wb, df):
    """
    Phase 1.2 — Tax Payments summary sheet.
    Rows use SUMIFS/COUNTIFS against m_Debit, m_CatFinal, m_SubCatFinal, m_TaxFlag.
    """
    ws = wb.create_sheet("🧾 Tax Payments")
    headers = ["Tax Type", "Amount", "Count"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "🧾 TAX PAYMENTS ANALYSIS", num_cols)
    _write_header_row(ws, 3, headers)

    rows_data = [
        ("All Tax Payments (Category)",      "=SUMIFS(m_Debit,m_CatFinal,\"Taxes\")",              "=COUNTIFS(m_CatFinal,\"Taxes\")"),
        ("  - Income Tax / Advance Tax",     "=SUMIFS(m_Debit,m_SubCatFinal,\"Tax Payment\")",     "=COUNTIFS(m_SubCatFinal,\"Tax Payment\")"),
        ("  - GST Payments",                 "=SUMIFS(m_Debit,m_SubCatFinal,\"GST Payment\")",     "=COUNTIFS(m_SubCatFinal,\"GST Payment\")"),
        ("Tax-Flagged Payments (Y)",         "=SUMIFS(m_Debit,m_TaxFlag,\"Y\")",                   "=COUNTIFS(m_TaxFlag,\"Y\")"),
        ("Tax-Flagged Payments (Yes)",       "=SUMIFS(m_Debit,m_TaxFlag,\"Yes\")",                 "=COUNTIFS(m_TaxFlag,\"Yes\")"),
    ]

    first_data_row = 4
    for i, (label, amount_formula, count_formula) in enumerate(rows_data):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18
        ws.cell(row=row, column=1, value=label).border = BORDER
        ws.cell(row=row, column=2, value=amount_formula).border = BORDER
        ws.cell(row=row, column=3, value=count_formula).border = BORDER

    # TOTAL row — sum category row + flag rows (rows 4, 7, 8 → indices 0, 3, 4)
    last_data_row = first_data_row + len(rows_data) - 1
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(row=total_row, column=1, value="TOTAL").border = BORDER
    ws.cell(row=total_row, column=2, value=f"=B{first_data_row}+B{first_data_row+3}+B{first_data_row+4}").border = BORDER
    ws.cell(row=total_row, column=3, value=f"=C{first_data_row}+C{first_data_row+3}+C{first_data_row+4}").border = BORDER
    _style_total_row(ws, total_row, num_cols)

    # Note row
    note_row = total_row + 1
    ws.row_dimensions[note_row].height = 18
    note_cell = ws.cell(row=note_row, column=1, value="Note: Flag-based rows (Y/Yes) will be 0 until Tax_Flag is set by CA.")
    note_cell.font = Font(name="Calibri", italic=True, color="7F7F7F")
    note_cell.border = BORDER
    for col in range(2, num_cols + 1):
        ws.cell(row=note_row, column=col).border = BORDER

    _apply_currency_format(ws, first_data_row, total_row, [2])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "B4"


def create_80c_tracker_sheet(wb, df):
    """
    Phase 1.2 — 80C Deduction Tracker sheet.
    Approximates 80C-eligible amounts using SUMIFS against m_CatFinal / m_SubCatFinal.
    """
    ws = wb.create_sheet("💼 80C Deduction Tracker")
    headers = ["80C Category", "Amount (₹)", "Note"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "💼 80C DEDUCTION TRACKER (APPROXIMATE)", num_cols)

    # Header note row (row 2 is spacer — add note in row 3, push headers to row 4)
    ws.merge_cells("A2:C2")
    note2_cell = ws.cell(row=2, column=1,
        value="Note: Individual 80C items share 'Investments' subcategory. Verify against challans, premium receipts, and CA-confirmed records.")
    note2_cell.font = Font(name="Calibri", italic=True, size=9, color="7F7F7F")
    note2_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, num_cols + 1):
        ws.cell(row=2, column=col).border = BORDER

    _write_header_row(ws, 3, headers)

    rows_data = [
        ("Life Insurance Premium",                  "=SUMIFS(m_Debit,m_CatFinal,\"Insurance\")",    "Sec 80C / 80D — Insurance premiums"),
        ("All Investments (PPF / ELSS / NPS / Sukanya)", "=SUMIFS(m_Debit,m_CatFinal,\"Investments\")", "Approx — all investment outflows (PPF, ELSS, NPS, Sukanya)"),
        ("School / Tuition Fees",                   0,                                              "No category mapping — manual entry required"),
        ("TOTAL 80C (Approximate)",                 None,                                           "Sum of above"),
    ]

    first_data_row = 4
    for i, (label, amount, note_text) in enumerate(rows_data):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18
        ws.cell(row=row, column=1, value=label).border = BORDER
        ws.cell(row=row, column=3, value=note_text).border = BORDER

        if amount is None:
            # TOTAL row — sum of data rows above (skip school/tuition which is static 0)
            total_formula = f"=B{first_data_row}+B{first_data_row+1}+B{first_data_row+2}"
            ws.cell(row=row, column=2, value=total_formula).border = BORDER
            _style_total_row(ws, row, num_cols)
        else:
            ws.cell(row=row, column=2, value=amount).border = BORDER

    # Footer note
    footer_row = first_data_row + len(rows_data)
    ws.row_dimensions[footer_row].height = 18
    footer_cell = ws.cell(row=footer_row, column=1,
        value="⚠️ These figures are approximate. Verify against investment challans, premium receipts, and CA-confirmed records.")
    footer_cell.font = Font(name="Calibri", italic=True, color="C00000")
    footer_cell.border = BORDER
    for col in range(2, num_cols + 1):
        ws.cell(row=footer_row, column=col).border = BORDER

    _apply_currency_format(ws, first_data_row, first_data_row + len(rows_data) - 1, [2])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.freeze_panes = "B4"


def create_capital_gain_sheet(wb, df):
    """
    Phase 1.2 — Capital Gain sheet.
    Uses SUMIFS/COUNTIFS against m_Credit, m_CatFinal, m_CGFlag.
    """
    ws = wb.create_sheet("📈 Capital Gain")
    headers = ["Capital Gain Type", "Credit Amount", "Count"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "📈 CAPITAL GAIN ANALYSIS", num_cols)
    _write_header_row(ws, 3, headers)

    rows_data = [
        ("Capital Gains (Category-based)",  "=SUMIFS(m_Credit,m_CatFinal,\"Capital Gains\")", "=COUNTIFS(m_CatFinal,\"Capital Gains\")"),
        ("Capital Gains (CG_Flag = Y)",     "=SUMIFS(m_Credit,m_CGFlag,\"Y\")",               "=COUNTIFS(m_CGFlag,\"Y\")"),
        ("Capital Gains (CG_Flag = Yes)",   "=SUMIFS(m_Credit,m_CGFlag,\"Yes\")",             "=COUNTIFS(m_CGFlag,\"Yes\")"),
    ]

    first_data_row = 4
    for i, (label, amount_formula, count_formula) in enumerate(rows_data):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18
        ws.cell(row=row, column=1, value=label).border = BORDER
        ws.cell(row=row, column=2, value=amount_formula).border = BORDER
        ws.cell(row=row, column=3, value=count_formula).border = BORDER

    # TOTAL row
    last_data_row = first_data_row + len(rows_data) - 1
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(row=total_row, column=1, value="TOTAL").border = BORDER
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})").border = BORDER
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})").border = BORDER
    _style_total_row(ws, total_row, num_cols)

    # Note row
    note_row = total_row + 1
    ws.row_dimensions[note_row].height = 18
    note_cell = ws.cell(row=note_row, column=1,
        value="Note: Flag-based rows will be 0 until CG_Flag is set. Detailed broker reconciliation in Phase 3.")
    note_cell.font = Font(name="Calibri", italic=True, color="7F7F7F")
    note_cell.border = BORDER
    for col in range(2, num_cols + 1):
        ws.cell(row=note_row, column=col).border = BORDER

    _apply_currency_format(ws, first_data_row, total_row, [2])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "B4"


def create_drawings_sheet(wb, df):
    """
    Phase 1.2 — Drawings sheet.
    Summarises personal/business drawings by category using SUMIFS/COUNTIFS.
    """
    ws = wb.create_sheet("🏧 Drawings")
    headers = ["Drawing Category", "Amount (₹)", "Count"]
    num_cols = len(headers)

    _write_formula_sheet_title(ws, "🏧 DRAWINGS & PERSONAL EXPENSE SUMMARY", num_cols)
    _write_header_row(ws, 3, headers)

    rows_data = [
        ("Tax Payments",             "=SUMIFS(m_Debit,m_CatFinal,\"Taxes\")",       "=COUNTIFS(m_CatFinal,\"Taxes\")"),
        ("GST Payments",             "=SUMIFS(m_Debit,m_SubCatFinal,\"GST Payment\")", "=COUNTIFS(m_SubCatFinal,\"GST Payment\")"),
        ("Utility Payments",         "=SUMIFS(m_Debit,m_CatFinal,\"Utilities\")",   "=COUNTIFS(m_CatFinal,\"Utilities\")"),
        ("Cash Withdrawals",         "=SUMIFS(m_Debit,m_CatFinal,\"Cash\")",        "=COUNTIFS(m_CatFinal,\"Cash\")"),
        ("Insurance Premium (Personal)", "=SUMIFS(m_Debit,m_CatFinal,\"Insurance\")", "=COUNTIFS(m_CatFinal,\"Insurance\")"),
        ("Food / Household",         "=SUMIFS(m_Debit,m_CatFinal,\"Food\")",        "=COUNTIFS(m_CatFinal,\"Food\")"),
        ("Shopping / Personal",      "=SUMIFS(m_Debit,m_CatFinal,\"Shopping\")",    "=COUNTIFS(m_CatFinal,\"Shopping\")"),
    ]

    first_data_row = 4
    for i, (label, amount_formula, count_formula) in enumerate(rows_data):
        row = first_data_row + i
        ws.row_dimensions[row].height = 18
        ws.cell(row=row, column=1, value=label).border = BORDER
        ws.cell(row=row, column=2, value=amount_formula).border = BORDER
        ws.cell(row=row, column=3, value=count_formula).border = BORDER

    # TOTAL row
    last_data_row = first_data_row + len(rows_data) - 1
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(row=total_row, column=1, value="TOTAL Drawings").border = BORDER
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})").border = BORDER
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})").border = BORDER
    _style_total_row(ws, total_row, num_cols)

    _apply_currency_format(ws, first_data_row, total_row, [2])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "B4"


def create_unclassified_sheet(wb, df):
    """
    Phase 1.2 — Unclassified Transactions sheet.
    Static filtered table of rows where Category_Final is Others/Miscellaneous
    or Confidence is Low/LOW.
    """
    ws = wb.create_sheet("❓ Unclassified Transactions")
    ws.sheet_view.showGridLines = False

    # Filter unclassified rows
    mask = (
        df["Category_Final"].isin(["Others", "Miscellaneous"]) |
        df["Confidence"].isin(["Low", "LOW"])
    )
    unclassified = df[mask].copy()

    display_cols = ["Transaction_ID", "Date", "Bank_Name", "Narration", "Debit", "Credit", "Confidence", "Category_Final"]
    # Keep only columns present
    display_cols = [c for c in display_cols if c in unclassified.columns]
    sub_df = unclassified[display_cols].copy()

    count = len(sub_df)
    total_debit = sub_df["Debit"].sum() if "Debit" in sub_df.columns else 0.0
    total_credit = sub_df["Credit"].sum() if "Credit" in sub_df.columns else 0.0

    num_cols = len(display_cols)
    last_col_letter = get_column_letter(num_cols)

    # Row 1: Title (merged)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=1, value="❓ UNCLASSIFIED TRANSACTIONS (Reflects engine classification at generation time)")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER

    # Row 2: Static summary
    ws.merge_cells(f"A2:{last_col_letter}2")
    summary_cell = ws.cell(row=2, column=1,
        value=f"Count: {count} | Total Debit: ₹{total_debit:,.2f} | Total Credit: ₹{total_credit:,.2f}")
    summary_cell.font = Font(name="Calibri", bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    summary_cell.border = BORDER
    ws.row_dimensions[2].height = 20
    for col in range(2, num_cols + 1):
        ws.cell(row=2, column=col).border = BORDER

    # Row 3: spacer
    ws.row_dimensions[3].height = 10
    for col in range(1, num_cols + 1):
        ws.cell(row=3, column=col).border = BORDER

    # Row 4: headers
    _write_header_row(ws, 4, display_cols)

    # Row 5+: data rows
    for r_idx, (_, row) in enumerate(sub_df.iterrows(), start=5):
        ws.row_dimensions[r_idx].height = 18
        for c_idx, col_name in enumerate(display_cols, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL

    if count == 0:
        ws.merge_cells(f"A5:{last_col_letter}5")
        empty_cell = ws.cell(row=5, column=1, value="No unclassified transactions found.")
        empty_cell.font = Font(name="Calibri", italic=True, color="7F7F7F")
        empty_cell.border = BORDER

    # Apply currency format to Debit and Credit columns
    if count > 0:
        debit_col = display_cols.index("Debit") + 1 if "Debit" in display_cols else None
        credit_col = display_cols.index("Credit") + 1 if "Credit" in display_cols else None
        currency_cols = [c for c in [debit_col, credit_col] if c is not None]
        if currency_cols:
            _apply_currency_format(ws, 5, 4 + count, currency_cols)

    # Column widths
    col_widths = {
        "Transaction_ID": 15, "Date": 14, "Bank_Name": 14,
        "Narration": 35, "Debit": 18, "Credit": 18,
        "Confidence": 14, "Category_Final": 20,
    }
    for c_idx, col_name in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "B5"


def create_ca_observations_sheet(wb, df, reconciliation_discrepancies: list):
    """
    Phase 1.2 — CA Observations sheet.
    Combines compliance flag audit (same 4 checks as Executive Summary) and
    reconciliation discrepancies from reconcile_all().
    """
    ws = wb.create_sheet("📌 CA Observations")
    ws.sheet_view.showGridLines = False

    num_cols = 3

    # Row 1: Title
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value="📌 CA OBSERVATIONS & RECONCILIATION DISCREPANCIES")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER

    # Row 2: spacer
    ws.row_dimensions[2].height = 10
    for col in range(1, num_cols + 1):
        ws.cell(row=2, column=col).border = BORDER

    # Section A: Compliance flags — headers in row 3
    _write_header_row(ws, 3, ["Flag", "Status", "Observations"])

    # Build compliance checklist (same 4 checks as executive summary)
    checklist = []

    total_cash_dep = df[(df["Credit"] > 0.0) & (df["Category"] == "Cash")]["Credit"].sum()
    if total_cash_dep > 200000.0:
        checklist.append(("SFT Savings Account Cash Deposit Limit (₹2 Lakh)", "🚩 RED FLAG",
            f"Total cash deposits of {format_inr(total_cash_dep)} exceed the ₹2 Lakh SFT threshold. Reportable under Rule 114E."))
    else:
        checklist.append(("SFT Savings Account Cash Deposit Limit (₹2 Lakh)", "✅ PASS",
            f"Total cash deposits of {format_inr(total_cash_dep)} are within safe SFT limits."))

    large_txns_count = df[(df["Credit"] >= 50000.0) | (df["Debit"] >= 50000.0)].shape[0]
    if large_txns_count > 0:
        checklist.append(("High Value Transactions (Sec 285BA / AIS Matching)", "⚠️ WARNING",
            f"Found {large_txns_count} transaction(s) >= ₹50,000. Ensure verification against AIS details."))
    else:
        checklist.append(("High Value Transactions (Sec 285BA / AIS Matching)", "✅ PASS",
            "No transactions >= ₹50,000 detected."))

    has_invest = df[df["Category"] == "Investments"].shape[0] > 0
    if has_invest:
        checklist.append(("Tax Saving Investment Audit (Sec 80C / 80D)", "✅ DETECTED",
            "Deduction flows found (SIP/PPF/NPS/Insurance). Verify investment challans for final computations."))
    else:
        checklist.append(("Tax Saving Investment Audit (Sec 80C / 80D)", "⚠️ NOT FOUND",
            "No common tax saving investments (Sec 80C/80D) were detected in the statements."))

    bounce_count = df[df["Category"] == "Bounces"].shape[0]
    if bounce_count > 0:
        checklist.append(("Cheque Bounce & ECS Return Audit (Sec 138 NI Act)", "🚩 RED FLAG",
            f"Detected {bounce_count} bounce/ECS return events. Review customer ledger balances."))
    else:
        checklist.append(("Cheque Bounce & ECS Return Audit (Sec 138 NI Act)", "✅ PASS",
            "No cheque bounces or mandate returns detected."))

    for i, (rule, status, obs) in enumerate(checklist, start=4):
        ws[f"A{i}"] = rule
        ws[f"B{i}"] = status
        ws[f"C{i}"] = obs
        ws.row_dimensions[i].height = 18
        for col in ["A", "B", "C"]:
            cell = ws[f"{col}{i}"]
            cell.border = BORDER
            if col == "B":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if "RED FLAG" in status:
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    cell.font = Font(name="Calibri", bold=True, color="C00000")
                elif "WARNING" in status or "NOT FOUND" in status:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    cell.font = Font(name="Calibri", bold=True, color="7F6000")
                else:
                    cell.fill = PatternFill("solid", fgColor="E2EFDA")
                    cell.font = Font(name="Calibri", bold=True, color="375623")
            elif i % 2 == 0:
                cell.fill = ALT_FILL

    # Row 8: spacer
    spacer_row = 4 + len(checklist)
    ws.row_dimensions[spacer_row].height = 10
    for col in range(1, num_cols + 1):
        ws.cell(row=spacer_row, column=col).border = BORDER

    # Row 9: Reconciliation Discrepancies sub-header
    recon_header_row = spacer_row + 1
    ws.merge_cells(f"A{recon_header_row}:C{recon_header_row}")
    recon_hdr_cell = ws.cell(row=recon_header_row, column=1, value="Reconciliation Discrepancies")
    recon_hdr_cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    recon_hdr_cell.fill = HDR_FILL
    recon_hdr_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[recon_header_row].height = 22
    for col in range(1, num_cols + 1):
        ws.cell(row=recon_header_row, column=col).fill = HDR_FILL
        ws.cell(row=recon_header_row, column=col).border = BORDER

    # Wide table for discrepancies — extend to 8 columns
    recon_num_cols = 8
    recon_col_headers = ["Bank", "Account", "Txn Index", "Prev Balance", "Expected", "Actual", "Difference", "Type"]

    # Row 10: discrepancy column headers
    recon_col_row = recon_header_row + 1
    ws.row_dimensions[recon_col_row].height = 22
    for c_idx, hdr in enumerate(recon_col_headers, start=1):
        cell = ws.cell(row=recon_col_row, column=c_idx, value=hdr)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Row 11+: discrepancy data
    recon_data_start = recon_col_row + 1
    if reconciliation_discrepancies:
        for r_idx, disc in enumerate(reconciliation_discrepancies, start=recon_data_start):
            ws.row_dimensions[r_idx].height = 18
            row_vals = [
                disc.get("bank", ""),
                disc.get("account", ""),
                disc.get("index", ""),
                disc.get("prev", 0.0),
                disc.get("expected", 0.0),
                disc.get("actual", 0.0),
                disc.get("diff", 0.0),
                disc.get("type", ""),
            ]
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = BORDER
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL
        # Currency format for balance columns (4,5,6,7)
        last_recon_row = recon_data_start + len(reconciliation_discrepancies) - 1
        _apply_currency_format(ws, recon_data_start, last_recon_row, [4, 5, 6, 7])
    else:
        ws.row_dimensions[recon_data_start].height = 18
        ws.merge_cells(f"A{recon_data_start}:H{recon_data_start}")
        no_disc_cell = ws.cell(row=recon_data_start, column=1, value="No reconciliation discrepancies found.")
        no_disc_cell.font = Font(name="Calibri", italic=True, color="375623")
        no_disc_cell.border = BORDER

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14

    ws.freeze_panes = "B4"


def create_inter_bank_transfers_sheet(wb, df):
    """🔁 Inter-Bank Transfers sheet — rows where Transfer_Group != ''."""
    ws = wb.create_sheet("🔁 Inter-Bank Transfers")
    ws.sheet_view.showGridLines = False

    # Filter to IBT rows
    if "Transfer_Group" in df.columns:
        ibt_df = df[df["Transfer_Group"] != ""].copy()
    else:
        ibt_df = pd.DataFrame()

    display_cols = ["Transfer_Group", "Date", "Bank_Name", "Account_Number", "Narration", "Debit", "Credit"]
    display_cols = [c for c in display_cols if c in df.columns]

    if not ibt_df.empty:
        ibt_df = ibt_df.sort_values("Transfer_Group")
        sub_df = ibt_df[display_cols].copy()
    else:
        sub_df = pd.DataFrame(columns=display_cols)

    count = len(sub_df)
    num_cols = len(display_cols)
    last_col_letter = get_column_letter(num_cols) if num_cols > 0 else "A"

    # Row 1: Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=1, value="🔁 INTER-BANK TRANSFERS")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER

    # Row 2: Count summary
    ws.merge_cells(f"A2:{last_col_letter}2")
    summary_cell = ws.cell(row=2, column=1, value=f"Total inter-bank transfer legs: {count}")
    summary_cell.font = Font(name="Calibri", bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    summary_cell.border = BORDER
    ws.row_dimensions[2].height = 20
    for col in range(2, num_cols + 1):
        ws.cell(row=2, column=col).border = BORDER

    # Row 3: Headers
    _write_header_row(ws, 3, display_cols)

    # Row 4+: Data
    for r_idx, (_, row) in enumerate(sub_df.iterrows(), start=4):
        ws.row_dimensions[r_idx].height = 18
        for c_idx, col_name in enumerate(display_cols, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL

    if count == 0:
        ws.merge_cells(f"A4:{last_col_letter}4")
        empty_cell = ws.cell(row=4, column=1, value="No inter-bank transfers detected.")
        empty_cell.font = Font(name="Calibri", italic=True, color="7F7F7F")
        empty_cell.border = BORDER

    # Currency format for Debit and Credit
    if count > 0:
        debit_col = display_cols.index("Debit") + 1 if "Debit" in display_cols else None
        credit_col = display_cols.index("Credit") + 1 if "Credit" in display_cols else None
        currency_cols = [c for c in [debit_col, credit_col] if c is not None]
        if currency_cols:
            _apply_currency_format(ws, 4, 3 + count, currency_cols)

    # Column widths
    col_widths = {"Transfer_Group": 14, "Date": 14, "Bank_Name": 16,
                  "Account_Number": 18, "Narration": 35, "Debit": 16, "Credit": 16}
    for c_idx, col_name in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "B4"


def create_loan_ledger_sheet(wb, df):
    """🏦 Loan Ledger sheet — output of build_loan_ledger."""
    ws = wb.create_sheet("🏦 Loan Ledger")
    ledger_df = build_loan_ledger(df)
    populate_and_format_table_sheet(ws, "🏦 LOAN LEDGER SUMMARY", ledger_df)


def create_emi_analysis_sheet(wb, df):
    """📉 EMI Analysis sheet — rows where Loan_Role in {'EMI', 'Repayment'}."""
    ws = wb.create_sheet("📉 EMI Analysis")
    ws.sheet_view.showGridLines = False

    if "Loan_Role" in df.columns:
        emi_df = df[df["Loan_Role"].isin({"EMI", "Repayment"})].copy()
    else:
        emi_df = pd.DataFrame()

    display_cols = ["Date", "Bank_Name", "Counterparty", "Narration", "Debit", "Loan_ID", "Loan_Status"]
    display_cols = [c for c in display_cols if c in df.columns]

    if not emi_df.empty:
        sub_df = emi_df[display_cols].copy()
    else:
        sub_df = pd.DataFrame(columns=display_cols)

    count = len(sub_df)
    num_cols = len(display_cols)
    last_col_letter = get_column_letter(num_cols) if num_cols > 0 else "A"

    # Row 1: Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=1, value="📉 EMI & LOAN REPAYMENT ANALYSIS")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER

    # Row 2: Count summary
    ws.merge_cells(f"A2:{last_col_letter}2")
    total_emi = sub_df["Debit"].sum() if "Debit" in sub_df.columns and count > 0 else 0.0
    summary_cell = ws.cell(row=2, column=1,
        value=f"EMI/Repayment rows: {count} | Total outflow: ₹{total_emi:,.2f}")
    summary_cell.font = Font(name="Calibri", bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    summary_cell.border = BORDER
    ws.row_dimensions[2].height = 20
    for col in range(2, num_cols + 1):
        ws.cell(row=2, column=col).border = BORDER

    # Row 3: Headers
    _write_header_row(ws, 3, display_cols)

    # Row 4+: Data
    for r_idx, (_, row) in enumerate(sub_df.iterrows(), start=4):
        ws.row_dimensions[r_idx].height = 18
        for c_idx, col_name in enumerate(display_cols, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL

    if count == 0:
        ws.merge_cells(f"A4:{last_col_letter}4")
        empty_cell = ws.cell(row=4, column=1, value="No EMI / loan repayment rows detected.")
        empty_cell.font = Font(name="Calibri", italic=True, color="7F7F7F")
        empty_cell.border = BORDER

    # Currency format for Debit column
    if count > 0 and "Debit" in display_cols:
        debit_col = display_cols.index("Debit") + 1
        _apply_currency_format(ws, 4, 3 + count, [debit_col])

    col_widths = {"Date": 14, "Bank_Name": 16, "Counterparty": 20,
                  "Narration": 35, "Debit": 16, "Loan_ID": 12, "Loan_Status": 16}
    for c_idx, col_name in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "B4"


def create_related_party_sheet(wb, df):
    """👥 Related Party sheet — counterparties appearing on both credit and debit, or Loan_Role in Receipt/Repayment."""
    ws = wb.create_sheet("👥 Related Party")
    ws.sheet_view.showGridLines = False

    display_cols = ["Counterparty", "Date", "Narration", "Debit", "Credit", "Loan_ID"]
    display_cols = [c for c in display_cols if c in df.columns]

    # Build related party mask
    rp_df = pd.DataFrame()
    if not df.empty:
        # Condition 1: Counterparty appears on both a credit and a debit
        cp_with_credit = set(df[df["Credit"] > 0]["Counterparty"].dropna().unique())
        cp_with_debit = set(df[df["Debit"] > 0]["Counterparty"].dropna().unique())
        both_sides = cp_with_credit & cp_with_debit
        # Remove blank/generic counterparties
        both_sides = {cp for cp in both_sides if cp.strip() not in ("", "N/A", "OTHERS")}
        mask_both = df["Counterparty"].isin(both_sides)

        # Condition 2: Loan_Role in Receipt or Repayment
        if "Loan_Role" in df.columns:
            mask_loan = df["Loan_Role"].isin({"Receipt", "Repayment"})
        else:
            mask_loan = pd.Series(False, index=df.index)

        rp_df = df[mask_both | mask_loan][display_cols].copy()

    count = len(rp_df)
    num_cols = len(display_cols)
    last_col_letter = get_column_letter(num_cols) if num_cols > 0 else "A"

    # Row 1: Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=1, value="👥 RELATED PARTY TRANSACTIONS")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).fill = HDR_FILL
        ws.cell(row=1, column=col).border = BORDER

    # Row 2: Count summary
    ws.merge_cells(f"A2:{last_col_letter}2")
    summary_cell = ws.cell(row=2, column=1, value=f"Related party rows: {count}")
    summary_cell.font = Font(name="Calibri", bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    summary_cell.border = BORDER
    ws.row_dimensions[2].height = 20
    for col in range(2, num_cols + 1):
        ws.cell(row=2, column=col).border = BORDER

    # Row 3: Headers
    _write_header_row(ws, 3, display_cols)

    # Row 4+: Data
    for r_idx, (_, row) in enumerate(rp_df.iterrows(), start=4):
        ws.row_dimensions[r_idx].height = 18
        for c_idx, col_name in enumerate(display_cols, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL

    if count == 0:
        ws.merge_cells(f"A4:{last_col_letter}4")
        empty_cell = ws.cell(row=4, column=1, value="No related party transactions detected.")
        empty_cell.font = Font(name="Calibri", italic=True, color="7F7F7F")
        empty_cell.border = BORDER

    # Currency format for Debit and Credit
    if count > 0:
        debit_col = display_cols.index("Debit") + 1 if "Debit" in display_cols else None
        credit_col = display_cols.index("Credit") + 1 if "Credit" in display_cols else None
        currency_cols = [c for c in [debit_col, credit_col] if c is not None]
        if currency_cols:
            _apply_currency_format(ws, 4, 3 + count, currency_cols)

    col_widths = {"Counterparty": 22, "Date": 14, "Narration": 35,
                  "Debit": 16, "Credit": 16, "Loan_ID": 12}
    for c_idx, col_name in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "B4"


def create_transaction_sheet(wb, sheet_name, df_bank):
    """Creates a standardized bank-specific transaction ledger sheet."""
    ws = wb.create_sheet(sheet_name)
    df_bank = df_bank.sort_values(by=["Date", "Balance"]).copy()

    display_cols = ["Date", "Account_Number", "Narration", "Chq_Ref", "Debit", "Credit", "Balance",
                    "Merchant_Name", "Transaction_Mode", "Category", "Sub_Category"]
    sub_df = df_bank[display_cols].copy()
    sub_df = sub_df.rename(columns={"Account_Number": "Account No"})

    title_text = sheet_name.replace("📒 ", "").upper()
    populate_and_format_table_sheet(ws, f"📒 {title_text}", sub_df)


def create_consolidated_master_sheet(wb, df: pd.DataFrame):
    """
    Phase 0 — Consolidated 'All Transactions' master sheet.

    Display columns chosen to give CAs full context + editable overrides.
    Original-evidence columns are LOCKED; editable override columns are UNLOCKED.
    Sheet protection is enabled (no password) so that locked cells cannot be
    accidentally overwritten, but the sheet can be unprotected at any time.
    """
    ws = wb.create_sheet("📋 All Transactions")
    df_sorted = df.sort_values(by=["Bank_Name", "Date", "txn_seq"]).copy()

    # Ordered display columns for the master sheet
    display_cols = [
        # Evidence (locked)
        "Transaction_ID",
        "Date",
        "Financial_Year",
        "Bank_Name",
        "Account_Number",
        "Statement_File_Name",
        "Sheet_Name",
        "Statement_Row_No",
        "Narration",
        "Chq_Ref",
        "Debit",
        "Credit",
        "Balance",
        "Transaction_Mode",
        "Merchant_Name",
        "Counterparty",
        "Confidence",
        "Match_Reason",
        # Engine outputs (locked — source of truth, CAs edit *_Final copies)
        "Category",
        "Sub_Category",
        # Transaction engine tags (locked — evidence)
        "Transfer_Group",
        "Loan_ID",
        "Loan_Role",
        "Loan_Status",
        # Editable overrides (unlocked)
        "Category_Final",
        "Sub_Category_Final",
        "GST_Flag",
        "Loan_Flag",
        "Tax_Flag",
        "CG_Flag",
        "Remarks",
    ]

    # Keep only columns present in df (graceful)
    display_cols = [c for c in display_cols if c in df_sorted.columns]
    sub_df = df_sorted[display_cols].copy()

    # Build the sheet layout (title, spacer, headers, data rows) via the standard helper
    populate_and_format_table_sheet(
        ws,
        "📋 CONSOLIDATED ALL TRANSACTIONS — MASTER LEDGER",
        sub_df,
        apply_table=True,
    )

    # -----------------------------------------------------------------
    # Phase 0: Apply cell-level protection
    # Header row = 3, data starts at row 4
    # -----------------------------------------------------------------
    locked_set = set(LOCKED_COLUMNS) | {"Category", "Sub_Category",
                                         "Transaction_Mode", "Merchant_Name",
                                         "Counterparty", "Confidence", "Match_Reason",
                                         "Chq_Ref"}
    unlocked_set = set(UNLOCKED_COLUMNS)

    # Map column name -> column index (1-based) in the sheet
    col_name_to_idx = {
        ws.cell(row=3, column=c).value: c
        for c in range(1, len(display_cols) + 1)
    }

    max_data_row = ws.max_row

    for col_name, col_idx in col_name_to_idx.items():
        is_unlocked = col_name in unlocked_set
        for row_idx in range(4, max_data_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.protection = Protection(locked=not is_unlocked)

    # Lock header and title rows (rows 1-3) — always locked
    for row_idx in range(1, 4):
        for col_idx in range(1, len(display_cols) + 1):
            ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=True)

    # Enable sheet protection (no password — easily unlockable)
    ws.protection = SheetProtection(sheet=True, password="")
    ws.protection.enable()

    # Freeze pane stays at B4 (set by populate_and_format_table_sheet)

    # Phase 1.1 — return mapping so build_ca_report can define named ranges
    return col_name_to_idx
