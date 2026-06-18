from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from ca_analyzer.core.config import config

FONT_FAMILY = config.styles.get("presentation", {}).get("font_family", "Calibri")

HDR_FILL = PatternFill("solid", fgColor=config.styles.get("presentation", {}).get("header_fill", "1F4E78"))
HDR_FONT = Font(name=FONT_FAMILY, bold=True, color=config.styles.get("presentation", {}).get("header_font_color", "FFFFFF"), size=11)

SUB_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_FONT = Font(name=FONT_FAMILY, bold=True, color="FFFFFF", size=10)

TITLE_FONT = Font(name=FONT_FAMILY, bold=True, size=14, color=config.styles.get("presentation", {}).get("header_fill", "1F4E78"))

ALT_FILL = PatternFill("solid", fgColor=config.styles.get("presentation", {}).get("alternate_row_fill", "F2F5F9"))

BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

CREDITS_FILL = PatternFill("solid", fgColor=config.styles.get("conditional_formatting", {}).get("credits_fill", "E2EFDA"))
CREDITS_FONT = Font(name=FONT_FAMILY, color=config.styles.get("conditional_formatting", {}).get("credits_font", "375623"))

DEBITS_FILL = PatternFill("solid", fgColor=config.styles.get("conditional_formatting", {}).get("debits_fill", "FCE4D6"))
DEBITS_FONT = Font(name=FONT_FAMILY, color=config.styles.get("conditional_formatting", {}).get("debits_font", "C65911"))

LOW_BAL_FILL = PatternFill("solid", fgColor=config.styles.get("conditional_formatting", {}).get("low_balance_fill", "FFF2CC"))
LOW_BAL_FONT = Font(name=FONT_FAMILY, color=config.styles.get("conditional_formatting", {}).get("low_font", "7F6000"))

NEG_CASH_FILL = PatternFill("solid", fgColor=config.styles.get("conditional_formatting", {}).get("negative_cashflow_fill", "C00000"))
NEG_CASH_FONT = Font(name=FONT_FAMILY, color=config.styles.get("conditional_formatting", {}).get("negative_cashflow_font", "FFFFFF"), bold=True)

HIGH_VAL_FILL = PatternFill("solid", fgColor=config.styles.get("conditional_formatting", {}).get("high_value_fill", "FFFFCC"))
HIGH_VAL_FONT = Font(name=FONT_FAMILY, color=config.styles.get("conditional_formatting", {}).get("high_value_font", "7F7F00"))

BANK_COLORS = config.styles.get("bank_colors", {})
CURRENCY_FORMAT = config.styles.get("presentation", {}).get("currency_format", "₹#,##0.00")
