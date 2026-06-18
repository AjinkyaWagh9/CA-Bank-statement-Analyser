import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import uuid
import pandas as pd
from ca_analyzer.presentation.styles import (
    HDR_FILL, HDR_FONT, ALT_FILL, BORDER, LEFT, CENTER, RIGHT, CURRENCY_FORMAT,
    CREDITS_FILL, CREDITS_FONT, DEBITS_FILL, DEBITS_FONT, LOW_BAL_FILL, LOW_BAL_FONT,
    HIGH_VAL_FILL, HIGH_VAL_FONT, BANK_COLORS
)

def clean_sheet_name(title: str) -> str:
    """Cleans a sheet title to be alphanumeric and underscores only for Excel Table name."""
    import re
    cleaned = re.sub(r"[^\w\s]", "", title)  # remove emojis, punctuation
    cleaned = cleaned.replace(" ", "_").strip("_")
    return cleaned

def populate_and_format_table_sheet(ws, title_text: str, df_data: pd.DataFrame, headers: list = None, apply_table: bool = True, header_row: int = 3):
    """
    Standard layout generator for tabular sheets:
    - Row 1: Merged Title cell (background #1F4E78, white bold font, size 14, height 25)
    - Row 2: Blank Spacer (height 15)
    - Row 3: Headers Row (background #1F4E78, white bold font, size 11, height 22)
    - Row 4+: Data Rows (height 18)
    - Dynamic freeze panes set to B4
    - Auto-fit column widths (minimum width constraints)
    - Column-name-based currency formatting (₹#,##0.00)
    - Bank-wise column coloring
    - Rule-based conditional formatting (avoiding blank cells)
    - Unique Excel Table wrapping (TableStyleMedium2)
    """
    ws.sheet_view.showGridLines = False
    
    max_c = max(df_data.shape[1], 1)
    last_col_letter = get_column_letter(max_c)
    
    # 1. Dynamically Merged Title Block in Row 1
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Ensure merged cells receive same background fill and border styles
    for col in range(1, max_c + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HDR_FILL
        cell.border = BORDER
        
    # 2. Blank Spacer in Row 2
    ws.row_dimensions[2].height = 15
    for col in range(1, max_c + 1):
        ws.cell(row=2, column=col).value = None
        ws.cell(row=2, column=col).border = BORDER
        
    # 3. Write Headers in Row 3
    ws.row_dimensions[header_row].height = 22
    columns_list = headers if headers is not None else list(df_data.columns)
    
    for col_idx, col_name in enumerate(columns_list, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = str(col_name) if col_name is not None and str(col_name).strip() != "" else f"Column_{col_idx}"
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        
    # 4. Write Data starting from Row 4
    for r_idx, (_, row) in enumerate(df_data.iterrows(), start=header_row + 1):
        ws.row_dimensions[r_idx].height = 18
        is_alt = (r_idx % 2 == 0)
        
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = BORDER
            
            # Apply manual striping ONLY if no Excel Table is applied
            if not apply_table and is_alt:
                cell.fill = ALT_FILL
            else:
                cell.fill = PatternFill(fill_type=None)
                
            # Clean values
            if pd.isna(val):
                cell.value = None
            else:
                if isinstance(val, pd.Timestamp):
                    cell.value = val.strftime("%Y-%m-%d")
                else:
                    cell.value = val

    # 5. Format Column Alignments, Widths, Currencies, and Colors
    min_widths = {
        "date": 15,
        "bank": 15,
        "account": 20,
        "narration": 40,
        "merchant": 30,
        "counterparty": 30,
        "category": 25,
        "subcategory": 25,
        "debit": 18,
        "credit": 18,
        "balance": 18,
        "amount": 18,
        "total": 18,
        "emi": 18,
        "inflow": 18,
        "outflow": 18,
    }
    
    max_r = ws.max_row
    
    for col in range(1, max_c + 1):
        col_letter = get_column_letter(col)
        header_val = str(ws.cell(row=header_row, column=col).value or "").lower()
        
        # Predefined minimum widths
        p_min = 12
        for key, w in min_widths.items():
            if key in header_val:
                p_min = w
                break
                
        # Length-based autofit
        longest = 0
        for r in range(header_row, max_r + 1):
            val_str = str(ws.cell(row=r, column=col).value or "")
            if len(val_str) > longest:
                longest = len(val_str)
                
        col_width = max(p_min, longest + 2)
        if "narration" in header_val:
            col_width = min(col_width, 60)
            
        ws.column_dimensions[col_letter].width = col_width
        
        # Cell formats & alignments
        for r in range(header_row + 1, max_r + 1):
            cell = ws.cell(row=r, column=col)
            val = cell.value
            
            # Format currencies case-insensitive
            is_currency = any(x in header_val for x in ["debit", "credit", "balance", "amount", "total", "emi", "inflow", "outflow"])
            if is_currency:
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = RIGHT
                if val is not None and str(val).strip() != "":
                    try:
                        cell.value = float(str(val).replace("₹", "").replace(",", "").strip())
                    except ValueError:
                        pass
            elif "date" in header_val:
                cell.alignment = CENTER
            elif any(x in header_val for x in ["count", "frequency"]):
                cell.alignment = CENTER
            elif "narration" in header_val:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            else:
                cell.alignment = LEFT
                
            # Bank color coding
            if "bank" in header_val and val is not None:
                bank_name = str(val).upper()
                bg_color = None
                for bkey, color in BANK_COLORS.items():
                    if bkey in bank_name:
                        bg_color = color
                        break
                if not bg_color:
                    bg_color = BANK_COLORS.get("UNKNOWN", "F2F2F2")
                cell.fill = PatternFill("solid", fgColor=bg_color)
                
            # Rule-based conditional formatting (avoiding blank cells)
            if val is not None and str(val).strip() != "":
                try:
                    num_val = float(str(val).replace("₹", "").replace(",", "").strip())
                except ValueError:
                    num_val = None
                    
                if num_val is not None:
                    # Credit Green (if Credit > 0)
                    if "credit" in header_val and num_val > 0.0:
                        cell.fill = CREDITS_FILL
                        cell.font = CREDITS_FONT
                    # Debit Red (if Debit > 0)
                    elif "debit" in header_val and num_val > 0.0:
                        cell.fill = DEBITS_FILL
                        cell.font = DEBITS_FONT
                    # Low Balance Orange (if Balance < 10000)
                    elif "balance" in header_val and num_val < 10000.0:
                        cell.fill = LOW_BAL_FILL
                        cell.font = LOW_BAL_FONT
                    # High Value Yellow (if is_currency and >= 50000)
                    if is_currency and num_val >= 50000.0:
                        cell.fill = HIGH_VAL_FILL
                        cell.font = HIGH_VAL_FONT

    # 6. Freeze pane below header
    ws.freeze_panes = "B4"
    
    # 7. Add Excel Table with unique name
    if apply_table and max_r > header_row and not df_data.empty:
        c_title = clean_sheet_name(ws.title)
        tbl_name = f"tbl_{c_title}_{uuid.uuid4().hex[:8]}"
        
        start_letter = get_column_letter(1)
        end_letter = get_column_letter(max_c)
        table_ref = f"{start_letter}{header_row}:{end_letter}{max_r}"
        
        # Mandatory assertions
        assert table_ref.startswith("A3"), f"Table ref '{table_ref}' must start with A3"
        assert ws.max_row >= 3
        assert not df_data.empty
        headers_checked = [ws.cell(row=header_row, column=c_idx).value for c_idx in range(1, max_c + 1)]
        assert all(h is not None and str(h).strip() != "" for h in headers_checked), f"Excel Tables require non-empty headers: {headers_checked}"
        assert len(headers_checked) == len(set(headers_checked)), f"Excel Tables require unique headers: {headers_checked}"
        
        tab = Table(displayName=tbl_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        try:
            ws.add_table(tab)
        except Exception:
            pass

def format_custom_table_range(ws, start_row: int, end_row: int, max_col: int, table_title_clean: str, header_row: int, apply_table: bool = True):
    """
    Helper to style a table range in a custom layout worksheet (e.g. multi-table).
    Enforces currency format, alignments, bank colors, and conditional formatting.
    Registers an Excel Table structure with unique naming.
    """
    min_widths = {
        "date": 15,
        "bank": 15,
        "account": 20,
        "narration": 40,
        "merchant": 30,
        "counterparty": 30,
        "category": 25,
        "subcategory": 25,
        "debit": 18,
        "credit": 18,
        "balance": 18,
        "amount": 18,
        "total": 18,
        "emi": 18,
        "inflow": 18,
        "outflow": 18,
    }
    
    # Format headers in header_row
    ws.row_dimensions[header_row].height = 22
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            cell.value = str(cell.value)
        else:
            cell.value = f"Column_{col}"
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        
    for r in range(header_row + 1, end_row + 1):
        ws.row_dimensions[r].height = 18
        is_alt = (r % 2 == 0)
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            
            if not apply_table and is_alt:
                cell.fill = ALT_FILL
            elif apply_table:
                cell.fill = PatternFill(fill_type=None)
                
            header_val = str(ws.cell(row=header_row, column=col).value or "").lower()
            val = cell.value
            
            # Format currencies case-insensitive
            is_currency = any(x in header_val for x in ["debit", "credit", "balance", "amount", "total", "emi", "inflow", "outflow"])
            if is_currency:
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = RIGHT
                if val is not None and str(val).strip() != "":
                    try:
                        cell.value = float(str(val).replace("₹", "").replace(",", "").strip())
                    except ValueError:
                        pass
            elif "date" in header_val:
                cell.alignment = CENTER
            elif any(x in header_val for x in ["count", "frequency"]):
                cell.alignment = CENTER
            elif "narration" in header_val:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            else:
                cell.alignment = LEFT
                
            # Bank color coding
            if "bank" in header_val and val is not None:
                bank_name = str(val).upper()
                bg_color = None
                for bkey, color in BANK_COLORS.items():
                    if bkey in bank_name:
                        bg_color = color
                        break
                if not bg_color:
                    bg_color = BANK_COLORS.get("UNKNOWN", "F2F2F2")
                cell.fill = PatternFill("solid", fgColor=bg_color)
                
            # Rule-based conditional formatting (avoiding blank cells)
            if val is not None and str(val).strip() != "":
                try:
                    num_val = float(str(val).replace("₹", "").replace(",", "").strip())
                except ValueError:
                    num_val = None
                    
                if num_val is not None:
                    # Credit Green (if Credit > 0)
                    if "credit" in header_val and num_val > 0.0:
                        cell.fill = CREDITS_FILL
                        cell.font = CREDITS_FONT
                    # Debit Red (if Debit > 0)
                    elif "debit" in header_val and num_val > 0.0:
                        cell.fill = DEBITS_FILL
                        cell.font = DEBITS_FONT
                    # Low Balance Orange (if Balance < 10000)
                    elif "balance" in header_val and num_val < 10000.0:
                        cell.fill = LOW_BAL_FILL
                        cell.font = LOW_BAL_FONT
                    # High Value Yellow (if is_currency and >= 50000)
                    if is_currency and num_val >= 50000.0:
                        cell.fill = HIGH_VAL_FILL
                        cell.font = HIGH_VAL_FONT

    # Auto-fit column widths dynamically
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        header_val = str(ws.cell(row=header_row, column=col).value or "").lower()
        
        p_min = 12
        for key, w in min_widths.items():
            if key in header_val:
                p_min = w
                break
                
        longest = 0
        for r in range(header_row, end_row + 1):
            val_str = str(ws.cell(row=r, column=col).value or "")
            if len(val_str) > longest:
                longest = len(val_str)
                
        col_width = max(p_min, longest + 2)
        if "narration" in header_val:
            col_width = min(col_width, 60)
            
        # Ensure column width is set if it is currently smaller
        current_w = ws.column_dimensions[col_letter].width or 0
        if col_width > current_w:
            ws.column_dimensions[col_letter].width = col_width
            
    # Add Table range
    if apply_table and end_row > header_row:
        tbl_name = f"tbl_{table_title_clean}_{uuid.uuid4().hex[:8]}"
        start_letter = get_column_letter(1)
        end_letter = get_column_letter(max_col)
        table_ref = f"{start_letter}{header_row}:{end_letter}{end_row}"
        
        # Mandatory assertions
        assert table_ref.startswith(f"A{header_row}"), f"Table range must start at A{header_row}, got: {table_ref}"
        assert ws.max_row >= header_row
        headers_checked = [ws.cell(row=header_row, column=c_idx).value for c_idx in range(1, max_col + 1)]
        assert all(h is not None and str(h).strip() != "" for h in headers_checked), f"Excel Tables require non-empty headers: {headers_checked}"
        assert len(headers_checked) == len(set(headers_checked)), f"Excel Tables require unique headers: {headers_checked}"
        
        tab = Table(displayName=tbl_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        try:
            ws.add_table(tab)
        except Exception:
            pass
