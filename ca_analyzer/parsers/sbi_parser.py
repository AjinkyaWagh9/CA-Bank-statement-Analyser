import openpyxl
import re
import pandas as pd
from datetime import datetime
from ca_analyzer.parsers.base_parser import BaseParser
from ca_analyzer.core.exceptions import ParserError
from ca_analyzer.core.utilities import parse_amount, parse_date

class SBIParser(BaseParser):
    def extract_metadata(self) -> dict:
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb.active
        
        person_name = "N/A"
        account_number = "N/A"
        ifsc = "SBIN0013238"
        start_date = "N/A"
        end_date = "N/A"
        
        for r in range(1, 25):
            col1 = str(ws.cell(row=r, column=1).value or "").strip()
            col2 = ws.cell(row=r, column=2).value
            col2_str = str(col2 or "").strip()
            
            if "Account Name" in col1:
                person_name = col2_str
            elif "Account Number" in col1:
                account_number = col2_str.replace("_", "").strip()
            elif "IFS Code" in col1:
                ifsc = col2_str.strip()
            elif "Start Date" in col1:
                if isinstance(col2, datetime):
                    start_date = col2.strftime("%d/%m/%Y")
                else:
                    start_date = col2_str
            elif "End Date" in col1:
                if isinstance(col2, datetime):
                    end_date = col2.strftime("%d/%m/%Y")
                else:
                    end_date = col2_str
                    
        return {
            "Person_Name": person_name,
            "Bank_Name": "SBI",
            "Account_Number": account_number,
            "IFSC": ifsc,
            "Statement_Start": start_date,
            "Statement_End": end_date
        }

    def extract_transactions(self) -> pd.DataFrame:
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb.active
        
        header_row_idx = None
        for r in range(1, 30):
            row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
            if "txn date" in row_vals and "description" in row_vals and "balance" in row_vals:
                header_row_idx = r
                break
                
        if header_row_idx is None:
            raise ParserError("Could not find transaction header row for SBI statement.")
            
        headers = [str(ws.cell(row=header_row_idx, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        
        rows = []
        for r in range(header_row_idx + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
            if all(x is None or str(x).strip() == "" for x in row_vals):
                continue
                
            val_0 = str(row_vals[0]).strip()
            if "total" in val_0.lower() or "statement" in val_0.lower() or "note" in val_0.lower():
                break
            rows.append(row_vals)
            
        df = pd.DataFrame(rows, columns=headers)
        df.columns = [c.strip() for c in df.columns]
        
        col_map = {
            "Txn Date": "raw_date",
            "Description": "raw_narration",
            "Ref No./Cheque No.": "raw_chq_ref",
            "Debit": "raw_debit",
            "Credit": "raw_credit",
            "Balance": "raw_balance"
        }
        
        df_renamed = pd.DataFrame()
        for excel_col, canonical_col in col_map.items():
            if excel_col in df.columns:
                df_renamed[canonical_col] = df[excel_col]
            else:
                df_renamed[canonical_col] = 0.0 if "debit" in canonical_col or "credit" in canonical_col or "balance" in canonical_col else ""
                
        return df_renamed
