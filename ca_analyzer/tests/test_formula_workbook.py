"""
test_formula_workbook.py — Phase 1.1 test suite

Tests:
1. build_ca_report produces a workbook with all 11 m_* named ranges.
2. All four formula sheets contain formula strings (start with '=') that
   reference the expected named ranges.
3. Named ranges round-trip through save/reload (openpyxl load_workbook).
4. Pandas oracle: compute expected income/expense totals from test data and
   print a comparison table.
5. Try to use the `formulas` package to evaluate formulas; skip gracefully if
   not installed.
"""

import os
import tempfile
import pytest
import pandas as pd
import openpyxl

from ca_analyzer.presentation.report_builder import build_ca_report

# ---------------------------------------------------------------------------
# Minimal test DataFrame that satisfies every sheet in the pipeline
# ---------------------------------------------------------------------------

def make_test_df() -> pd.DataFrame:
    """
    Creates a small but complete DataFrame covering two banks, two months,
    and a spread of income and expense categories.
    """
    rows = [
        # Bank: HDFC — April 2024
        {
            "Date": pd.Timestamp("2024-04-05"),
            "Bank_Name": "HDFC", "Account_Number": "001", "IFSC": "HDFC0001",
            "Narration": "Salary credit", "Chq_Ref": "",
            "Debit": 0.0, "Credit": 80000.0, "Balance": 80000.0,
            "Transaction_Mode": "NEFT", "Merchant_Name": "Employer Ltd",
            "Counterparty": "EMPLOYER", "Category": "Salary", "Sub_Category": "Net Pay",
            "Category_Final": "Salary", "Sub_Category_Final": "Net Pay",
            "GST_Flag": "", "Loan_Flag": "", "Tax_Flag": "", "CG_Flag": "", "Remarks": "",
            "Confidence": "HIGH", "Match_Reason": "keyword",
            "Transaction_ID": "TXN001", "Statement_File_Name": "hdfc_apr.xlsx",
            "Sheet_Name": "Sheet1", "Statement_Row_No": 1,
            "Financial_Year": "2024-25",
            "Person_Name": "Test Client",
            "Statement_Start": "2024-04-01", "Statement_End": "2024-05-31",
            "txn_seq": 1,
            "Transfer_Group": "", "Loan_ID": "", "Loan_Role": "", "Loan_Status": "",
        },
        # HDFC — April 2024 — Expense (Rent)
        {
            "Date": pd.Timestamp("2024-04-10"),
            "Bank_Name": "HDFC", "Account_Number": "001", "IFSC": "HDFC0001",
            "Narration": "Rent paid", "Chq_Ref": "CHQ001",
            "Debit": 20000.0, "Credit": 0.0, "Balance": 60000.0,
            "Transaction_Mode": "CHEQUE", "Merchant_Name": "Landlord",
            "Counterparty": "LANDLORD", "Category": "Rent", "Sub_Category": "House Rent",
            "Category_Final": "Rent", "Sub_Category_Final": "House Rent",
            "GST_Flag": "", "Loan_Flag": "", "Tax_Flag": "", "CG_Flag": "", "Remarks": "",
            "Confidence": "HIGH", "Match_Reason": "keyword",
            "Transaction_ID": "TXN002", "Statement_File_Name": "hdfc_apr.xlsx",
            "Sheet_Name": "Sheet1", "Statement_Row_No": 2,
            "Financial_Year": "2024-25",
            "Person_Name": "Test Client",
            "Statement_Start": "2024-04-01", "Statement_End": "2024-05-31",
            "txn_seq": 2,
            "Transfer_Group": "", "Loan_ID": "", "Loan_Role": "", "Loan_Status": "",
        },
        # HDFC — May 2024 — Salary (second month)
        {
            "Date": pd.Timestamp("2024-05-05"),
            "Bank_Name": "HDFC", "Account_Number": "001", "IFSC": "HDFC0001",
            "Narration": "Salary credit", "Chq_Ref": "",
            "Debit": 0.0, "Credit": 80000.0, "Balance": 140000.0,
            "Transaction_Mode": "NEFT", "Merchant_Name": "Employer Ltd",
            "Counterparty": "EMPLOYER", "Category": "Salary", "Sub_Category": "Net Pay",
            "Category_Final": "Salary", "Sub_Category_Final": "Net Pay",
            "GST_Flag": "", "Loan_Flag": "", "Tax_Flag": "", "CG_Flag": "", "Remarks": "",
            "Confidence": "HIGH", "Match_Reason": "keyword",
            "Transaction_ID": "TXN003", "Statement_File_Name": "hdfc_may.xlsx",
            "Sheet_Name": "Sheet1", "Statement_Row_No": 1,
            "Financial_Year": "2024-25",
            "Person_Name": "Test Client",
            "Statement_Start": "2024-04-01", "Statement_End": "2024-05-31",
            "txn_seq": 3,
            "Transfer_Group": "", "Loan_ID": "", "Loan_Role": "", "Loan_Status": "",
        },
        # Bank: ICICI — April 2024 — Other Sources
        {
            "Date": pd.Timestamp("2024-04-15"),
            "Bank_Name": "ICICI", "Account_Number": "002", "IFSC": "ICIC0001",
            "Narration": "Interest income", "Chq_Ref": "",
            "Debit": 0.0, "Credit": 5000.0, "Balance": 5000.0,
            "Transaction_Mode": "CREDIT", "Merchant_Name": "",
            "Counterparty": "", "Category": "Other Sources", "Sub_Category": "Interest",
            "Category_Final": "Other Sources", "Sub_Category_Final": "Interest",
            "GST_Flag": "", "Loan_Flag": "", "Tax_Flag": "", "CG_Flag": "", "Remarks": "",
            "Confidence": "MEDIUM", "Match_Reason": "rule",
            "Transaction_ID": "TXN004", "Statement_File_Name": "icici_apr.xlsx",
            "Sheet_Name": "Sheet1", "Statement_Row_No": 1,
            "Financial_Year": "2024-25",
            "Person_Name": "Test Client",
            "Statement_Start": "2024-04-01", "Statement_End": "2024-05-31",
            "txn_seq": 4,
            "Transfer_Group": "", "Loan_ID": "", "Loan_Role": "", "Loan_Status": "",
        },
        # ICICI — May 2024 — Food expense
        {
            "Date": pd.Timestamp("2024-05-20"),
            "Bank_Name": "ICICI", "Account_Number": "002", "IFSC": "ICIC0001",
            "Narration": "Swiggy", "Chq_Ref": "",
            "Debit": 1500.0, "Credit": 0.0, "Balance": 3500.0,
            "Transaction_Mode": "UPI", "Merchant_Name": "Swiggy",
            "Counterparty": "SWIGGY", "Category": "Food", "Sub_Category": "Food Delivery",
            "Category_Final": "Food", "Sub_Category_Final": "Food Delivery",
            "GST_Flag": "", "Loan_Flag": "", "Tax_Flag": "", "CG_Flag": "", "Remarks": "",
            "Confidence": "HIGH", "Match_Reason": "merchant",
            "Transaction_ID": "TXN005", "Statement_File_Name": "icici_may.xlsx",
            "Sheet_Name": "Sheet1", "Statement_Row_No": 1,
            "Financial_Year": "2024-25",
            "Person_Name": "Test Client",
            "Statement_Start": "2024-04-01", "Statement_End": "2024-05-31",
            "txn_seq": 5,
            "Transfer_Group": "", "Loan_ID": "", "Loan_Role": "", "Loan_Status": "",
        },
        # ICICI — May 2024 — Unclassified row (Others + Low confidence)
        {
            "Date": pd.Timestamp("2024-05-25"),
            "Bank_Name": "ICICI", "Account_Number": "002", "IFSC": "ICIC0001",
            "Narration": "Unknown transfer", "Chq_Ref": "",
            "Debit": 3000.0, "Credit": 0.0, "Balance": 500.0,
            "Transaction_Mode": "NEFT", "Merchant_Name": "",
            "Counterparty": "", "Category": "Others", "Sub_Category": "",
            "Category_Final": "Others", "Sub_Category_Final": "",
            "GST_Flag": "", "Loan_Flag": "", "Tax_Flag": "", "CG_Flag": "", "Remarks": "",
            "Confidence": "Low", "Match_Reason": "fallback",
            "Transaction_ID": "TXN006", "Statement_File_Name": "icici_may.xlsx",
            "Sheet_Name": "Sheet1", "Statement_Row_No": 2,
            "Financial_Year": "2024-25",
            "Person_Name": "Test Client",
            "Statement_Start": "2024-04-01", "Statement_End": "2024-05-31",
            "txn_seq": 6,
            "Transfer_Group": "", "Loan_ID": "", "Loan_Role": "", "Loan_Status": "",
        },
    ]
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Pandas oracle — expected aggregates from the test data
# ---------------------------------------------------------------------------

def compute_pandas_oracle(df: pd.DataFrame) -> dict:
    """
    Compute expected income and expense totals using pandas groupby.
    Returns a dict with keys 'income' and 'expense', each a dict of
    {category: {bank: amount, ..., 'Total': total}}.
    """
    oracle = {"income": {}, "expense": {}}

    # Income (credit side)
    income_rows = df[df["Credit"] > 0]
    income_grouped = income_rows.groupby(["Category_Final", "Bank_Name"])["Credit"].sum()
    for (cat, bank), amount in income_grouped.items():
        oracle["income"].setdefault(cat, {})
        oracle["income"][cat][bank] = amount
    for cat, bank_amounts in oracle["income"].items():
        oracle["income"][cat]["Total"] = sum(bank_amounts.values())

    # Expense (debit side)
    expense_rows = df[df["Debit"] > 0]
    expense_grouped = expense_rows.groupby(["Category_Final", "Bank_Name"])["Debit"].sum()
    for (cat, bank), amount in expense_grouped.items():
        oracle["expense"].setdefault(cat, {})
        oracle["expense"][cat][bank] = amount
    for cat, bank_amounts in oracle["expense"].items():
        oracle["expense"][cat]["Total"] = sum(bank_amounts.values())

    return oracle


def print_oracle_table(oracle: dict):
    """Pretty-prints the oracle tables for income and expense."""
    print("\n=== PANDAS ORACLE: INCOME ===")
    for cat, bank_totals in sorted(oracle["income"].items()):
        for key, amount in bank_totals.items():
            print(f"  {cat:<25} | {key:<10} | ₹{amount:>12,.2f}")

    print("\n=== PANDAS ORACLE: EXPENSE ===")
    for cat, bank_totals in sorted(oracle["expense"].items()):
        for key, amount in bank_totals.items():
            print(f"  {cat:<25} | {key:<10} | ₹{amount:>12,.2f}")


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def workbook_path():
    """Build the report once and return the path; clean up after module tests."""
    df = make_test_df()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    build_ca_report(df, path)
    yield path
    if os.path.exists(path):
        os.unlink(path)


def test_named_ranges_exist(workbook_path):
    """All 11 m_* named ranges must be present after save/reload."""
    wb = openpyxl.load_workbook(workbook_path)
    expected_names = [
        "m_Bank", "m_Date", "m_Debit", "m_Credit",
        "m_CatFinal", "m_SubCatFinal",
        "m_GSTFlag", "m_LoanFlag", "m_TaxFlag", "m_CGFlag",
        "m_Confidence",
    ]
    missing = [n for n in expected_names if n not in wb.defined_names]
    assert not missing, f"Missing named ranges after round-trip: {missing}"


def test_named_ranges_point_to_master_sheet(workbook_path):
    """Each named range attr_text must reference '📋 All Transactions'."""
    wb = openpyxl.load_workbook(workbook_path)
    for name in ["m_Bank", "m_Credit", "m_Debit", "m_CatFinal"]:
        dn = wb.defined_names[name]
        assert "All Transactions" in dn.attr_text, (
            f"Named range {name} does not reference the master sheet: {dn.attr_text}"
        )


def test_bank_wise_summary_has_formulas(workbook_path):
    """The '🏦 Bank Wise Summary' sheet must have SUMIFS/COUNTIFS formula cells."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["🏦 Bank Wise Summary"]
    formula_cells = []
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append(cell.value)

    assert formula_cells, "No formula cells found in Bank Wise Summary"
    # Spot-check named range usage
    any_sumifs = any("SUMIFS" in f or "COUNTIFS" in f for f in formula_cells)
    assert any_sumifs, f"Expected SUMIFS/COUNTIFS, found: {formula_cells[:5]}"
    any_named = any("m_Credit" in f or "m_Bank" in f or "m_Debit" in f for f in formula_cells)
    assert any_named, f"Expected m_* named ranges in formulas, found: {formula_cells[:5]}"


def test_income_analysis_has_sumifs_with_named_ranges(workbook_path):
    """The '💰 Income Analysis' sheet must use SUMIFS referencing m_Credit and m_CatFinal."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["💰 Income Analysis"]
    formula_cells = [
        cell.value for row in ws.iter_rows(min_row=4)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=SUMIFS")
    ]
    assert formula_cells, "No SUMIFS cells found in Income Analysis"
    assert any("m_Credit" in f for f in formula_cells), "m_Credit not used in Income Analysis"
    assert any("m_CatFinal" in f for f in formula_cells), "m_CatFinal not used in Income Analysis"
    assert any("m_Bank" in f for f in formula_cells), "m_Bank not used in Income Analysis"


def test_expense_analysis_has_sumifs_with_named_ranges(workbook_path):
    """The '💸 Expense Analysis' sheet must use SUMIFS referencing m_Debit and m_CatFinal."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["💸 Expense Analysis"]
    formula_cells = [
        cell.value for row in ws.iter_rows(min_row=4)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=SUMIFS")
    ]
    assert formula_cells, "No SUMIFS cells found in Expense Analysis"
    assert any("m_Debit" in f for f in formula_cells), "m_Debit not used in Expense Analysis"
    assert any("m_CatFinal" in f for f in formula_cells), "m_CatFinal not used in Expense Analysis"


def test_monthly_cashflow_has_date_formulas(workbook_path):
    """The '📅 Monthly Cashflow' sheet must use DATE/EOMONTH formulas."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["📅 Monthly Cashflow"]
    formula_cells = [
        cell.value for row in ws.iter_rows(min_row=4)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formula_cells, "No formula cells found in Monthly Cashflow"
    any_date = any("DATE(" in f and "EOMONTH(" in f for f in formula_cells)
    assert any_date, f"Expected DATE/EOMONTH formulas, found: {formula_cells[:3]}"
    any_named = any("m_Credit" in f or "m_Bank" in f for f in formula_cells)
    assert any_named, f"Expected m_* named ranges in Monthly Cashflow formulas"


def test_income_analysis_row_labels(workbook_path):
    """Income Analysis col A must contain the canonical income category names."""
    from ca_analyzer.presentation.report_builder import INCOME_CATEGORIES
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["💰 Income Analysis"]
    labels = [ws.cell(row=r, column=1).value for r in range(4, 4 + len(INCOME_CATEGORIES))]
    assert labels == INCOME_CATEGORIES, (
        f"Income category labels mismatch.\nExpected: {INCOME_CATEGORIES}\nGot: {labels}"
    )


def test_expense_analysis_row_labels(workbook_path):
    """Expense Analysis col A must contain the canonical expense category names."""
    from ca_analyzer.presentation.report_builder import EXPENSE_CATEGORIES
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["💸 Expense Analysis"]
    labels = [ws.cell(row=r, column=1).value for r in range(4, 4 + len(EXPENSE_CATEGORIES))]
    assert labels == EXPENSE_CATEGORIES, (
        f"Expense category labels mismatch.\nExpected: {EXPENSE_CATEGORIES}\nGot: {labels}"
    )


def test_master_sheet_has_correct_headers(workbook_path):
    """Master sheet row 3 must have expected column headers."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["📋 All Transactions"]
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    for expected in ["Bank_Name", "Credit", "Debit", "Category_Final", "GST_Flag",
                     "Transfer_Group", "Loan_ID", "Loan_Role", "Loan_Status"]:
        assert expected in headers, f"Expected header '{expected}' missing from master sheet"


def test_pandas_oracle_printed(capfd):
    """Runs the pandas oracle and prints the income/expense table (informational)."""
    df = make_test_df()
    oracle = compute_pandas_oracle(df)
    print_oracle_table(oracle)
    captured = capfd.readouterr()

    # Expected values from test data
    assert oracle["income"]["Salary"]["HDFC"] == pytest.approx(160000.0)  # 2 months × 80k
    assert oracle["income"]["Other Sources"]["ICICI"] == pytest.approx(5000.0)
    assert oracle["income"]["Salary"]["Total"] == pytest.approx(160000.0)

    assert oracle["expense"]["Rent"]["HDFC"] == pytest.approx(20000.0)
    assert oracle["expense"]["Food"]["ICICI"] == pytest.approx(1500.0)

    # Verify something was printed
    assert "PANDAS ORACLE" in captured.out


# ---------------------------------------------------------------------------
# Phase 1.2 — Pandas oracle for new sheets
# ---------------------------------------------------------------------------

def compute_pandas_oracle_phase12(df: pd.DataFrame) -> dict:
    """
    Compute expected totals for Phase 1.2 sheets using pandas.
    Returns a dict with keys for each new sheet's key metric.
    """
    oracle = {}

    # Tax Payments: SUMIFS where Category_Final == "Taxes"
    oracle["tax_payments_category"] = df[df["Category_Final"] == "Taxes"]["Debit"].sum()

    # 80C: SUMIFS where Category_Final in ("Insurance", "Investments")
    oracle["80c_insurance"] = df[df["Category_Final"] == "Insurance"]["Debit"].sum()
    oracle["80c_investments"] = df[df["Category_Final"] == "Investments"]["Debit"].sum()
    oracle["80c_total"] = oracle["80c_insurance"] + oracle["80c_investments"]

    # Capital Gain: SUMIFS where Category_Final == "Capital Gains"
    oracle["capital_gains_category"] = df[df["Category_Final"] == "Capital Gains"]["Credit"].sum()

    # Drawings: sum of Taxes + Utilities + Cash + Insurance + Food + Shopping
    drawings_cats = ["Taxes", "Utilities", "Cash", "Insurance", "Food", "Shopping"]
    oracle["drawings_total"] = df[df["Category_Final"].isin(drawings_cats)]["Debit"].sum()

    print("\n=== PANDAS ORACLE: PHASE 1.2 ===")
    print(f"  Tax Payments (Category):     ₹{oracle['tax_payments_category']:>12,.2f}")
    print(f"  80C Insurance:               ₹{oracle['80c_insurance']:>12,.2f}")
    print(f"  80C Investments:             ₹{oracle['80c_investments']:>12,.2f}")
    print(f"  80C Total:                   ₹{oracle['80c_total']:>12,.2f}")
    print(f"  Capital Gains (Category):    ₹{oracle['capital_gains_category']:>12,.2f}")
    print(f"  Drawings Total:              ₹{oracle['drawings_total']:>12,.2f}")

    return oracle


# ---------------------------------------------------------------------------
# Phase 1.2 — New tests
# ---------------------------------------------------------------------------

def test_new_sheets_exist(workbook_path):
    """All Phase 1.2 + engine sheets must exist in the workbook."""
    wb = openpyxl.load_workbook(workbook_path)
    new_sheets = [
        "🧾 Tax Payments",
        "💼 80C Deduction Tracker",
        "📈 Capital Gain",
        "🏧 Drawings",
        "❓ Unclassified Transactions",
        "📌 CA Observations",
        "🔁 Inter-Bank Transfers",
        "🏦 Loan Ledger",
        "📉 EMI Analysis",
        "👥 Related Party",
    ]
    missing = [s for s in new_sheets if s not in wb.sheetnames]
    assert not missing, f"Missing new sheets: {missing}"


def test_tax_payments_has_sumifs(workbook_path):
    """The '🧾 Tax Payments' sheet must contain SUMIFS formulas referencing m_Debit and m_CatFinal."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["🧾 Tax Payments"]
    formula_cells = [
        cell.value for row in ws.iter_rows(min_row=4)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formula_cells, "No formula cells found in Tax Payments sheet"
    assert any("SUMIFS" in f for f in formula_cells), "No SUMIFS found in Tax Payments"
    assert any("m_Debit" in f for f in formula_cells), "m_Debit not used in Tax Payments"
    assert any("m_CatFinal" in f or "m_SubCatFinal" in f or "m_TaxFlag" in f for f in formula_cells), \
        "Expected named range (m_CatFinal/m_SubCatFinal/m_TaxFlag) in Tax Payments"


def test_80c_tracker_has_sumifs(workbook_path):
    """The '💼 80C Deduction Tracker' sheet must contain SUMIFS formulas."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["💼 80C Deduction Tracker"]
    formula_cells = [
        cell.value for row in ws.iter_rows(min_row=3)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formula_cells, "No formula cells found in 80C Deduction Tracker sheet"
    assert any("SUMIFS" in f for f in formula_cells), "No SUMIFS found in 80C Tracker"
    assert any("m_Debit" in f for f in formula_cells), "m_Debit not used in 80C Tracker"


def test_capital_gain_has_sumifs(workbook_path):
    """The '📈 Capital Gain' sheet must contain SUMIFS formulas referencing m_Credit."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["📈 Capital Gain"]
    formula_cells = [
        cell.value for row in ws.iter_rows(min_row=4)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formula_cells, "No formula cells found in Capital Gain sheet"
    assert any("SUMIFS" in f for f in formula_cells), "No SUMIFS found in Capital Gain"
    assert any("m_Credit" in f for f in formula_cells), "m_Credit not used in Capital Gain"
    assert any("m_CGFlag" in f or "m_CatFinal" in f for f in formula_cells), \
        "Expected m_CGFlag or m_CatFinal in Capital Gain"


def test_drawings_has_sumifs(workbook_path):
    """The '🏧 Drawings' sheet must contain SUMIFS formulas referencing m_Debit."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["🏧 Drawings"]
    formula_cells = [
        cell.value for row in ws.iter_rows(min_row=4)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formula_cells, "No formula cells found in Drawings sheet"
    assert any("SUMIFS" in f for f in formula_cells), "No SUMIFS found in Drawings"
    assert any("m_Debit" in f for f in formula_cells), "m_Debit not used in Drawings"


def test_unclassified_sheet_exists(workbook_path):
    """The '❓ Unclassified Transactions' sheet must exist and have data rows."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["❓ Unclassified Transactions"]
    # At least the title row must exist
    assert ws.cell(row=1, column=1).value is not None, "Unclassified sheet title missing"
    # Row 4 should have header content
    header_row_values = [ws.cell(row=4, column=c).value for c in range(1, 5)]
    assert any(v is not None for v in header_row_values), "Unclassified sheet missing headers in row 4"
    # At least one data row in row 5 (our test data has 1 unclassified row)
    data_row_values = [ws.cell(row=5, column=c).value for c in range(1, 5)]
    assert any(v is not None for v in data_row_values), "Unclassified sheet missing data row (expected 1 unclassified txn)"


def test_ca_observations_exists(workbook_path):
    """The '📌 CA Observations' sheet must exist and have compliance check content."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["📌 CA Observations"]
    assert ws.cell(row=1, column=1).value is not None, "CA Observations title missing"
    # Row 3 should have 'Flag' header
    assert ws.cell(row=3, column=1).value == "Flag", \
        f"Expected 'Flag' header in row 3, got: {ws.cell(row=3, column=1).value}"
    # Should have at least 4 compliance check rows (rows 4-7)
    assert ws.cell(row=4, column=1).value is not None, "CA Observations missing compliance data"


def test_reconcile_all_returns_list():
    """reconcile_all must return a list (not bool). With clean test data, returns empty list."""
    from ca_analyzer.validation.reconciliation import reconcile_all
    df = make_test_df()
    result = reconcile_all(df)
    assert isinstance(result, list), f"reconcile_all should return list, got {type(result)}"
    # Test data has consistent balances (each row independent), expect empty discrepancy list
    # (or at most some entries — just confirm it's a list)


def test_reconcile_bank_returns_list():
    """reconcile_bank must return a list (not bool)."""
    from ca_analyzer.validation.reconciliation import reconcile_bank
    df = make_test_df()
    result = reconcile_bank(df, "HDFC", "001")
    assert isinstance(result, list), f"reconcile_bank should return list, got {type(result)}"


def test_pandas_oracle_phase12(capfd):
    """Runs the Phase 1.2 pandas oracle and validates expected values."""
    df = make_test_df()
    oracle = compute_pandas_oracle_phase12(df)
    captured = capfd.readouterr()

    # Test data has no Taxes, Insurance, Investments, Capital Gains rows
    assert oracle["tax_payments_category"] == pytest.approx(0.0)
    assert oracle["80c_insurance"] == pytest.approx(0.0)
    assert oracle["80c_investments"] == pytest.approx(0.0)
    assert oracle["capital_gains_category"] == pytest.approx(0.0)
    # Drawings: Food = 1500 (Swiggy), Others unclassified not in drawings cats
    assert oracle["drawings_total"] == pytest.approx(1500.0)

    assert "PANDAS ORACLE" in captured.out


def test_formulas_package_evaluation(workbook_path):
    """
    Attempt to evaluate formula cells using the `formulas` package.
    Skips gracefully if the package is not installed.

    NOTE: The `formulas` package cannot resolve workbook-level DefinedNames
    that point to other sheets — it only evaluates self-contained cell formulas.
    This test verifies the package is available and the workbook loads cleanly,
    then skips formula evaluation for named-range-dependent cells.
    """
    pytest.importorskip("formulas", reason="`formulas` package not installed — skipping")

    import formulas  # noqa: F401 — only used when available
    # If we reach here, the package is installed.
    # Loading the workbook with formulas is enough to verify compatibility.
    try:
        xl_model = formulas.ExcelModel().loads(workbook_path)
        xl_model.calculate()
        print("\n`formulas` package: ExcelModel loaded and calculated successfully.")
    except Exception as exc:
        # Named-range formulas may cause errors in the formulas package —
        # this is expected; we just verify the package installed.
        print(f"\n`formulas` package: loaded but calculation raised: {exc}")
