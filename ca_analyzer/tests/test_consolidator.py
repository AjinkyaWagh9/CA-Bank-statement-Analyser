import os
import openpyxl
from ca_analyzer.consolidator import consolidate_statements

def test_consolidate_pipeline():
    files = [
        "Bank Statements/Unprocessed/SANJAY JINDAL HDFC BANK STATEMENT FY 2024-25.xls",
        "Bank Statements/Unprocessed/SANJAY JINDAL ICICI BANK FY 2024-25.xls",
        "Bank Statements/Unprocessed/SANJAY JINDAL SBI 2024-25.xlsx"
    ]
    
    files_present = [f for f in files if os.path.exists(f)]
    if len(files_present) == 3:
        output_xlsx = "Consolidated_CA_Analysis_Test.xlsx"
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)
            
        consolidate_statements(files_present, output_xlsx)
        assert os.path.exists(output_xlsx)
        assert os.path.getsize(output_xlsx) > 0
        
        # Load workbook and perform robust verification audits
        wb = openpyxl.load_workbook(output_xlsx)
        
        # 1. Properties Metadata verification
        assert wb.properties.creator == "CA Financial Intelligence Analyzer"
        assert wb.properties.title == "Consolidated Bank Statement Analysis"
        assert wb.properties.subject == "Multi-Bank Financial Analysis Report"
        
        # 2. Sheet Order verification
        expected_sheets = [
            "📋 Cover Sheet", "📝 Executive Summary", "🏦 Bank Wise Summary", "📅 Monthly Cashflow",
            "💰 Income Analysis", "💸 Expense Analysis", "💵 Cash Deposit Analysis", "🚨 High Value Transactions",
            "🏠 Loan Analysis", "🧾 GST Analysis", "🧾 TDS Analysis", "📈 Investment Analysis",
            "⚠️ Risk Flags", "📒 HDFC Transactions", "📒 ICICI Transactions", "📒 SBI Transactions"
        ]
        assert list(wb.sheetnames) == expected_sheets
        
        # 3. Standard Tabular Sheets layout and tables check
        standard_tabulars = [
            s for s in expected_sheets 
            if s not in ["📋 Cover Sheet", "📝 Executive Summary", "💵 Cash Deposit Analysis"]
        ]
        
        for sname in standard_tabulars:
            ws = wb[sname]
            
            # Row 1 Title Block checking
            assert ws["A1"].value is not None, f"Missing title on {sname}"
            # Check cell fill matches header fill (ends with 1F4E78)
            assert str(ws["A1"].fill.start_color.rgb).endswith("1F4E78"), f"Title color mismatched on {sname}"
            
            # Row 2 Spacer checking
            for col in range(1, ws.max_column + 1):
                assert ws.cell(row=2, column=col).value is None, f"Spacer cell in row 2 contains data on {sname}"
                
            # Row 3 Headers row checking
            assert ws.cell(row=3, column=1).value is not None, f"Missing header row 3 on {sname}"
            assert str(ws.cell(row=3, column=1).fill.start_color.rgb).endswith("1F4E78")
            
            # Dynamic freeze panes B4 checking
            assert ws.freeze_panes == "B4", f"Incorrect freeze pane on {sname}"
            
            # Excel Table checking (displayName and styles)
            if ws.max_row > 3:
                assert len(ws.tables) == 1, f"Missing Excel Table on {sname}"
                tbl = list(ws.tables.values())[0]
                assert tbl.displayName.startswith("tbl_"), f"Table name '{tbl.displayName}' does not start with tbl_ on {sname}"
            else:
                assert len(ws.tables) == 0, f"Table exists on empty sheet {sname}"
            
            # Unique headers check and no Unnamed columns
            seen_headers = set()
            for col_idx in range(1, ws.max_column + 1):
                hdr = ws.cell(row=3, column=col_idx).value
                assert hdr is not None, f"Header cell column {col_idx} is None on {sname}"
                assert "Unnamed" not in str(hdr), f"Unnamed header column {col_idx} on {sname}"
                assert str(hdr) not in seen_headers, f"Duplicate header '{hdr}' on {sname}"
                seen_headers.add(str(hdr))
                
            # Currency cells formatting verification
            for r in range(4, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    hdr_val = str(ws.cell(row=3, column=col_idx).value).lower()
                    cell = ws.cell(row=r, column=col_idx)
                    val = cell.value
                    
                    is_curr_col = any(x in hdr_val for x in ["debit", "credit", "balance", "amount", "total", "emi", "inflow", "outflow"])
                    if is_curr_col and val is not None and str(val).strip() != "":
                        assert cell.number_format == "₹#,##0.00", f"Cell {cell.coordinate} on {sname} missing Rupee format"
                        assert isinstance(val, (int, float)), f"Cell {cell.coordinate} on {sname} is not numeric float"
                        
        # 5. Multi-table Cash Deposit sheet verification
        cash_ws = wb["💵 Cash Deposit Analysis"]
        assert cash_ws.freeze_panes == "B4"
        assert len(cash_ws.tables) == 1, f"Expected exactly 1 Excel Table in Cash Deposit sheet, got {len(cash_ws.tables)}"
        
        # 6. Global Table Range starting index verification (startswith "A3")
        for sname in wb.sheetnames:
            ws = wb[sname]
            for tbl in ws.tables.values():
                assert tbl.ref.startswith("A3"), f"Table '{tbl.displayName}' in '{sname}' must start with A3: {tbl.ref}"
            
        # Clean up test output
        wb.close()
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)
